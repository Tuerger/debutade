from __future__ import annotations

from datetime import datetime, date
import getpass
import json
import logging
import os
import sys
from typing import Any

from flask import Flask, jsonify, redirect, render_template, request
from openpyxl import load_workbook


if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        import io

        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.getenv(
    "DEBUTADE_CONFIG",
    os.path.abspath(os.path.join(SCRIPT_DIR, "..", "config.json")),
)


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def parse_amount(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)

    raw = normalize_text(value)
    if not raw:
        return None

    raw = raw.replace("€", "").replace(" ", "")
    if raw.count(",") == 1 and raw.count(".") > 1:
        raw = raw.replace(".", "")
    raw = raw.replace(".", "").replace(",", ".")

    try:
        return float(raw)
    except ValueError:
        return None


def parse_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if value is None:
        return None

    raw = normalize_text(value)
    if not raw:
        return None

    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def normalize_side(value: Any) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""

    compact = raw.lower().replace(" ", "")
    if compact in {"af", "uit", "uitgave", "uitgaven"}:
        return "Af"
    if compact in {"bij", "in", "inkomst", "inkomsten"}:
        return "Bij"
    return raw


def load_config(config_path: str) -> dict[str, Any]:
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Configuratiebestand niet gevonden: {config_path}")

    with open(config_path, "r", encoding="utf-8") as config_file:
        root_config = json.load(config_file)

    shared = root_config.get("shared", {})
    bank = root_config.get("bankrekening", {})
    kas = root_config.get("kasboek", {})
    begroting = root_config.get("begroting", {})

    grootboek_directory = shared.get("grootboek_directory", "")

    def build_path(file_name: str) -> str:
        name = normalize_text(file_name)
        if not name:
            return ""
        if os.path.isabs(name):
            return name
        if grootboek_directory:
            return os.path.join(grootboek_directory, name)
        return name

    return {
        "grootboek_directory": grootboek_directory,
        "begroting_excel_path": build_path(begroting.get("excel_file_name", "")),
        "begroting_sheet_name": begroting.get("excel_sheet_name", "Begroting"),
        "bank_excel_path": build_path(
            bank.get("excel_file_name") or shared.get("bank_excel_file_name") or ""
        ),
        "bank_sheet_name": bank.get("excel_sheet_name", "Bankrekening"),
        "kas_excel_path": build_path(kas.get("excel_file_name", "")),
        "kas_sheet_name": kas.get("excel_sheet_name", "Kas"),
        "log_directory": shared.get("log_directory", os.path.join(SCRIPT_DIR, "logs")),
        "log_level": shared.get("log_level", "INFO"),
        "main_app_url": os.getenv("MAIN_APP_URL", "").strip(),
    }


try:
    config = load_config(CONFIG_PATH)
except (FileNotFoundError, KeyError, json.JSONDecodeError) as exc:
    print(f"WAARSCHUWING: {exc}")
    config = {
        "grootboek_directory": "",
        "begroting_excel_path": "",
        "begroting_sheet_name": "Begroting",
        "bank_excel_path": "",
        "bank_sheet_name": "Bankrekening",
        "kas_excel_path": "",
        "kas_sheet_name": "Kas",
        "log_directory": os.path.join(SCRIPT_DIR, "logs"),
        "log_level": "INFO",
        "main_app_url": os.getenv("MAIN_APP_URL", "").strip(),
    }


LOG_DIRECTORY = config["log_directory"]
LOG_LEVEL = config["log_level"]
MAIN_APP_URL = config["main_app_url"]

app = Flask(
    __name__,
    template_folder=os.path.join(SCRIPT_DIR, "templates"),
    static_folder=os.path.join(SCRIPT_DIR, "static"),
)
app.config["TEMPLATES_AUTO_RELOAD"] = True


def map_headers(sheet) -> dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    header_map: dict[str, int] = {}
    if not header_row:
        return header_map

    for index, cell_value in enumerate(header_row):
        key = normalize_text(cell_value).lower()
        if key and key not in header_map:
            header_map[key] = index
    return header_map


def get_value(
    row: tuple[Any, ...],
    header_map: dict[str, int],
    header_names: tuple[str, ...],
    fallback_index: int | None = None,
) -> Any:
    for header_name in header_names:
        idx = header_map.get(header_name.lower())
        if idx is not None and idx < len(row):
            return row[idx]
    if fallback_index is not None and fallback_index < len(row):
        return row[fallback_index]
    return None


def parse_financial_rows(file_path: str, sheet_name: str, fallback_tag_idx: int | None, fallback_af_bij_idx: int | None, fallback_amount_idx: int | None) -> tuple[list[dict[str, Any]], str | None]:
    if not file_path or not os.path.exists(file_path):
        return [], f"Bestand niet gevonden: {file_path}"

    workbook = None
    rows: list[dict[str, Any]] = []
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            return [], f"Sheet '{sheet_name}' niet gevonden in {file_path}"

        sheet = workbook[sheet_name]
        header_map = map_headers(sheet)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            tag = normalize_text(
                get_value(
                    row,
                    header_map,
                    ("tag", "grootboek", "categorie", "rekening"),
                    fallback_tag_idx,
                )
            )
            af_bij = normalize_side(
                get_value(
                    row,
                    header_map,
                    ("af bij", "af/bij", "type"),
                    fallback_af_bij_idx,
                )
            )
            amount = parse_amount(
                get_value(
                    row,
                    header_map,
                    ("bedrag (eur)", "bedrag", "amount", "begroot bedrag", "budget"),
                    fallback_amount_idx,
                )
            )

            if af_bij not in {"Af", "Bij"} or amount is None:
                continue

            if not tag:
                tag = "(Geen tag)"

            rows.append(
                {
                    "tag": tag,
                    "af_bij": af_bij,
                    "amount": round(float(amount), 2),
                    "date": parse_date(
                        get_value(row, header_map, ("datum",), 0)
                    ),
                }
            )

        return rows, None
    except Exception as exc:
        logging.error("Fout bij lezen van %s / %s: %s", file_path, sheet_name, exc)
        return [], f"Kon bestand niet verwerken: {file_path} ({exc})"
    finally:
        if workbook:
            workbook.close()


def parse_budget_rows(file_path: str, sheet_name: str) -> tuple[list[dict[str, Any]], str | None]:
    if not file_path or not os.path.exists(file_path):
        return [], f"Bestand niet gevonden: {file_path}"

    workbook = None
    rows: list[dict[str, Any]] = []
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            return [], f"Sheet '{sheet_name}' niet gevonden in {file_path}"

        sheet = workbook[sheet_name]
        header_map = map_headers(sheet)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            hoofdcategorie = normalize_text(
                get_value(
                    row,
                    header_map,
                    ("hoofdcategorie", "hoofd categorie", "hoofdcat", "categorie"),
                )
            ) or "(Geen hoofdcategorie)"

            subcategorie = normalize_text(
                get_value(
                    row,
                    header_map,
                    ("subcategorie", "sub categorie", "subcat", "sub-categorie"),
                )
            )

            tag = normalize_text(
                get_value(
                    row,
                    header_map,
                    ("tag", "grootboek", "rekening"),
                    11,
                )
            ) or "(Geen tag)"

            af_bij = normalize_side(
                get_value(
                    row,
                    header_map,
                    ("af bij", "af/bij", "bij/ af", "bij/af", "type"),
                    5,
                )
            )

            budget_amount = parse_amount(
                get_value(
                    row,
                    header_map,
                    ("bedrag begroot", "begroot bedrag", "budget", "bedrag"),
                    6,
                )
            )

            if af_bij not in {"Af", "Bij"} or budget_amount is None:
                continue

            rows.append(
                {
                    "hoofdcategorie": hoofdcategorie,
                    "subcategorie": subcategorie,
                    "tag": tag,
                    "af_bij": af_bij,
                    "budget": round(float(budget_amount), 2),
                }
            )

        return rows, None
    except Exception as exc:
        logging.error("Fout bij lezen van begroting %s / %s: %s", file_path, sheet_name, exc)
        return [], f"Kon begroting niet verwerken: {file_path} ({exc})"
    finally:
        if workbook:
            workbook.close()


def summarize_by_side(rows: list[dict[str, Any]]) -> dict[str, dict[str, float]]:
    result = {"Bij": {}, "Af": {}}

    for item in rows:
        side = item["af_bij"]
        tag = item["tag"]
        amount = item["amount"]
        result[side][tag] = round(result[side].get(tag, 0.0) + amount, 2)

    return result


def summarize_actual_by_tag(rows: list[dict[str, Any]]) -> dict[str, float]:
    result: dict[str, float] = {}
    for item in rows:
        tag = item.get("tag") or "(Geen tag)"
        amount = float(item.get("amount") or 0.0)
        result[tag] = round(result.get(tag, 0.0) + amount, 2)
    return result


def build_rows_for_ui(
    budget_rows: list[dict[str, Any]],
    actual_by_tag: dict[str, float],
    include_actual: bool,
) -> list[dict[str, Any]]:
    result = []

    for item in budget_rows:
        budget_value = round(float(item.get("budget") or 0.0), 2)
        actual_value = round(actual_by_tag.get(item.get("tag") or "", 0.0), 2)
        result.append(
            {
                "hoofdcategorie": item.get("hoofdcategorie") or "(Geen hoofdcategorie)",
                "subcategorie": item.get("subcategorie") or "",
                "tag": item.get("tag") or "(Geen tag)",
                "budget": budget_value,
                "actual": actual_value if include_actual else None,
                "difference": round(actual_value - budget_value, 2) if include_actual else None,
            }
        )

    return result


def load_budget_and_actual(include_actual: bool) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []

    budget_rows, budget_warning = parse_budget_rows(
        config.get("begroting_excel_path", ""),
        config.get("begroting_sheet_name", "Begroting"),
    )
    if budget_warning:
        warnings.append(budget_warning)

    budget_rows_bij = [row for row in budget_rows if row.get("af_bij") == "Bij"]
    budget_rows_af = [row for row in budget_rows if row.get("af_bij") == "Af"]

    actual_tag_bij: dict[str, float] = {}
    actual_tag_af: dict[str, float] = {}
    if include_actual:
        bank_rows, bank_warning = parse_financial_rows(
            config.get("bank_excel_path", ""),
            config.get("bank_sheet_name", "Bankrekening"),
            11,
            5,
            6,
        )
        if bank_warning:
            warnings.append(bank_warning)

        kas_rows, kas_warning = parse_financial_rows(
            config.get("kas_excel_path", ""),
            config.get("kas_sheet_name", "Kas"),
            11,
            5,
            6,
        )
        if kas_warning:
            warnings.append(kas_warning)

        actual_sides = summarize_by_side(bank_rows + kas_rows)
        actual_tag_bij = actual_sides["Bij"]
        actual_tag_af = actual_sides["Af"]

    inkomsten_rows = build_rows_for_ui(budget_rows_bij, actual_tag_bij, include_actual)
    uitgaven_rows = build_rows_for_ui(budget_rows_af, actual_tag_af, include_actual)

    totals = {
        "budget_inkomsten": round(sum(item["budget"] for item in inkomsten_rows), 2),
        "budget_uitgaven": round(sum(item["budget"] for item in uitgaven_rows), 2),
        "actual_inkomsten": round(sum((item["actual"] or 0.0) for item in inkomsten_rows), 2) if include_actual else None,
        "actual_uitgaven": round(sum((item["actual"] or 0.0) for item in uitgaven_rows), 2) if include_actual else None,
    }

    if include_actual:
        totals["budget_saldo"] = round(totals["budget_inkomsten"] - totals["budget_uitgaven"], 2)
        totals["actual_saldo"] = round((totals["actual_inkomsten"] or 0.0) - (totals["actual_uitgaven"] or 0.0), 2)
    else:
        totals["budget_saldo"] = round(totals["budget_inkomsten"] - totals["budget_uitgaven"], 2)
        totals["actual_saldo"] = None

    payload = {
        "inkomsten": inkomsten_rows,
        "uitgaven": uitgaven_rows,
        "totals": totals,
        "meta": {
            "include_actual": include_actual,
            "begroting_file": config.get("begroting_excel_path", ""),
            "begroting_sheet": config.get("begroting_sheet_name", "Begroting"),
            "bank_file": config.get("bank_excel_path", ""),
            "bank_sheet": config.get("bank_sheet_name", "Bankrekening"),
            "kas_file": config.get("kas_excel_path", ""),
            "kas_sheet": config.get("kas_sheet_name", "Kas"),
        },
    }
    return payload, warnings


@app.before_request
def log_request() -> None:
    logging.info("REQUEST %s %s %s", request.remote_addr, request.method, request.path)


@app.route("/")
def index():
    return render_template(
        "index.html",
        current_date=datetime.now().strftime("%d-%m-%Y"),
        current_user=getpass.getuser(),
        main_app_url=MAIN_APP_URL,
    )


@app.route("/api/begroting-data")
def begroting_data():
    include_actual = request.args.get("actual", "0") == "1"
    payload, warnings = load_budget_and_actual(include_actual)
    payload["warnings"] = warnings
    return jsonify(payload)


@app.route("/settings")
def settings():
    if MAIN_APP_URL:
        return redirect(f"{MAIN_APP_URL}/settings")
    return jsonify({"success": False, "message": "Instellingen zijn alleen beschikbaar via de hoofdapp."}), 403


@app.route("/quit", methods=["POST"])
def quit_application():
    try:
        user = getpass.getuser()
        logging.info("APPLICATIE AFGESLOTEN | Gebruiker: %s", user)
        logging.info("=" * 70)

        response = jsonify({"success": True, "message": "Applicatie sluit af"})

        def shutdown_server() -> None:
            import time

            time.sleep(1)
            logging.info("Flask server wordt beeindigd...")
            os._exit(0)

        import threading

        shutdown_thread = threading.Thread(target=shutdown_server, daemon=True)
        shutdown_thread.start()

        return response, 200
    except Exception as exc:
        logging.error("Fout bij afsluiten applicatie: %s", str(exc))
        return jsonify({"success": False, "message": f"Fout: {str(exc)}"}), 500


if __name__ == "__main__":
    if not os.path.exists(LOG_DIRECTORY):
        try:
            os.makedirs(LOG_DIRECTORY)
        except Exception as exc:
            print(f"FOUT: Kan log directory niet aanmaken: {LOG_DIRECTORY}")
            print(f"Details: {str(exc)}")
            exit(1)

    log_file_path = os.path.join(LOG_DIRECTORY, "begroting_webapp_log.txt")
    logging.basicConfig(
        filename=log_file_path,
        level=getattr(logging, str(LOG_LEVEL).upper(), logging.INFO),
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    logging.info("=" * 70)
    logging.info("BEGROTING START")
    logging.info("Begroting bestand: %s", config.get("begroting_excel_path", ""))
    logging.info("Bank bestand: %s", config.get("bank_excel_path", ""))
    logging.info("Kas bestand: %s", config.get("kas_excel_path", ""))
    logging.info("=" * 70)

    port = int(os.getenv("DEBUTADE_APP_PORT", "5004"))
    app.run(debug=False, host="127.0.0.1", port=port, use_reloader=False)
