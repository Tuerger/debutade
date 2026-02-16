import os
import sys
import time
import socket
import logging
import subprocess
import configparser
import json
from datetime import datetime

from flask import Flask, jsonify, render_template, redirect, request, url_for
from openpyxl import load_workbook

try:
    import psutil
except Exception:
    psutil = None


def get_app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def load_config(config_path):
    config = configparser.ConfigParser(interpolation=None)
    config["logging"] = {
        "level": "INFO",
        "file": "debutade.log",
        "format": "%(asctime)s [%(levelname)s] %(message)s",
    }
    try:
        config.read(config_path, encoding="utf-8")
    except Exception as exc:
        print(f"Kon configuratie niet laden: {exc}")
    return config


def setup_logging(config, app_dir):
    section = config["logging"] if "logging" in config else {}
    level_name = section.get("level", "INFO").upper()
    level = logging._nameToLevel.get(level_name, logging.INFO)

    log_file = section.get("file", "debutade.log")
    if not os.path.isabs(log_file):
        log_file = os.path.join(app_dir, log_file)

    log_dir = os.path.dirname(log_file)
    if log_dir and not os.path.exists(log_dir):
        try:
            os.makedirs(log_dir, exist_ok=True)
        except Exception as exc:
            print(f"WAARSCHUWING: Kon log directory niet aanmaken: {exc}")
            log_file = os.path.join(app_dir, "debutade.log")

    log_format = section.get("format", "%(asctime)s [%(levelname)s] %(message)s")
    logging.basicConfig(
        level=level,
        format=log_format,
        filename=log_file,
        filemode="a",
        force=True,
    )


APP_DIR = get_app_dir()
CONFIG_PATH = os.path.join(APP_DIR, "start-debutade.config")
config = load_config(CONFIG_PATH)
setup_logging(config, APP_DIR)

CONFIG_JSON_PATH = os.path.join(APP_DIR, "config.json")

MAIN_HOST = os.getenv("MAIN_APP_HOST", "127.0.0.1")
if MAIN_HOST in {"0.0.0.0", "::", "[::]"}:
    MAIN_HOST = "127.0.0.1"
MAIN_PORT = int(os.getenv("MAIN_APP_PORT", "5003"))
MAIN_APP_URL = f"http://{MAIN_HOST}:{MAIN_PORT}"
SUBAPP_PORT = int(os.getenv("SUBAPP_PORT", "5004"))
SUBAPP_START_TIMEOUT = int(os.getenv("SUBAPP_START_TIMEOUT", "60"))


def load_main_config():
    if not os.path.exists(CONFIG_JSON_PATH):
        return None, f"config.json niet gevonden: {CONFIG_JSON_PATH}"
    try:
        with open(CONFIG_JSON_PATH, "r", encoding="utf-8") as handle:
            return json.load(handle), None
    except Exception as exc:
        return None, f"Kon config.json niet laden: {exc}"


def save_main_config(config_data):
    try:
        with open(CONFIG_JSON_PATH, "w", encoding="utf-8") as handle:
            json.dump(config_data, handle, indent=4)
        return None
    except Exception as exc:
        return f"Opslaan mislukt: {exc}"


def split_lines(value):
    return [line.strip() for line in (value or "").splitlines() if line.strip()]


def validate_workbook_tabs(file_path, required_sheets):
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        return [f"Kon Excel bestand niet openen: {file_path} ({exc})"]
    try:
        sheet_names = set(wb.sheetnames)
        missing = [name for name in required_sheets if name not in sheet_names]
        if missing:
            return [
                f"Excel bestand mist tabblad(en): {', '.join(missing)} ({file_path})"
            ]
        return []
    finally:
        wb.close()


def validate_main_config(config_data):
    errors = []
    shared = config_data.get("shared", {})
    shared_bank_name = shared.get("bank_excel_file_name")

    grootboek_dir = shared.get("grootboek_directory")
    if not grootboek_dir:
        errors.append("Grootboek directory is verplicht.")
    elif not os.path.isdir(grootboek_dir):
        errors.append(f"Grootboek directory bestaat niet: {grootboek_dir}")

    for key, label in (
        ("backup_directory", "Backup directory"),
        ("log_directory", "Log directory"),
        ("resources", "Resources directory"),
    ):
        value = shared.get(key)
        if not value:
            errors.append(f"{label} is verplicht.")
        elif not os.path.isdir(value):
            errors.append(f"{label} bestaat niet: {value}")

    bank = config_data.get("bankrekening", {})
    kas = config_data.get("kasboek", {})
    bon = config_data.get("bontoevoegen", {})
    showreport = config_data.get("showreport", {})
    contributie = config_data.get("contributie", {})

    def build_excel_path(file_name):
        if not grootboek_dir or not file_name:
            return None
        return os.path.join(grootboek_dir, file_name)

    bank_file_name = bank.get("excel_file_name") or shared_bank_name
    if not bank_file_name:
        errors.append("Bank Excel bestandsnaam (gedeeld) is verplicht.")
    bank_path = build_excel_path(bank_file_name)
    if bank_path and not os.path.exists(bank_path):
        errors.append(f"Bankrekening Excel bestand niet gevonden: {bank_path}")

    bank_sheets = bank.get("required_sheets") or []
    if bank_path and bank_sheets:
        errors.extend(validate_workbook_tabs(bank_path, bank_sheets))
    elif bank_path and not bank_sheets:
        errors.append("Bankrekening vereiste sheets zijn verplicht.")

    kas_file_name = kas.get("excel_file_name")
    if not kas_file_name:
        errors.append("Kasboek Excel bestandsnaam is verplicht.")
    kas_path = build_excel_path(kas_file_name)
    if kas_path and not os.path.exists(kas_path):
        errors.append(f"Kasboek Excel bestand niet gevonden: {kas_path}")

    kas_sheet = kas.get("excel_sheet_name")
    if not kas_sheet:
        errors.append("Kasboek Excel sheet naam is verplicht.")
    elif kas_path:
        errors.extend(validate_workbook_tabs(kas_path, [kas_sheet]))

    bon_bank_name = bon.get("bank_excel_file_name") or shared_bank_name
    if not bon_bank_name:
        errors.append("Bon Toevoegen bank Excel bestandsnaam (gedeeld) is verplicht.")
    bon_bank_path = build_excel_path(bon_bank_name)
    if bon_bank_path and not os.path.exists(bon_bank_path):
        errors.append(f"Bon Toevoegen bank Excel bestand niet gevonden: {bon_bank_path}")

    bon_kas_name = bon.get("kas_excel_file_name")
    if not bon_kas_name:
        errors.append("Bon Toevoegen kas Excel bestandsnaam is verplicht.")
    bon_kas_path = build_excel_path(bon_kas_name)
    if bon_kas_path and not os.path.exists(bon_kas_path):
        errors.append(f"Bon Toevoegen kas Excel bestand niet gevonden: {bon_kas_path}")

    report_url = (showreport.get("report_url") or "").strip()
    if report_url and not (report_url.startswith("http://") or report_url.startswith("https://")):
        errors.append("Showreport URL moet starten met http:// of https://")

    contributie_file = contributie.get("ledenbestand_path")
    if not contributie_file:
        errors.append("Contributie ledenbestand pad is verplicht.")
    elif not os.path.exists(contributie_file):
        errors.append(f"Contributie ledenbestand niet gevonden: {contributie_file}")

    contributie_personen = contributie.get("leden_sheet_personen")
    contributie_betaald = contributie.get("leden_sheet_betaald")
    if not contributie_personen:
        errors.append("Contributie tab personen is verplicht.")
    if not contributie_betaald:
        errors.append("Contributie tab betaald is verplicht.")
    if contributie_file and contributie_personen and contributie_betaald:
        errors.extend(validate_workbook_tabs(
            contributie_file,
            [contributie_personen, contributie_betaald],
        ))

    contributie_bank_name = contributie.get("bank_excel_file_name") or shared_bank_name
    if not contributie_bank_name:
        errors.append("Contributie bank Excel bestandsnaam (gedeeld) is verplicht.")
    contributie_bank_path = build_excel_path(contributie_bank_name)
    if contributie_bank_path and not os.path.exists(contributie_bank_path):
        errors.append(f"Contributie bank Excel bestand niet gevonden: {contributie_bank_path}")

    contributie_bank_sheet = contributie.get("bank_sheet_name")
    if not contributie_bank_sheet:
        errors.append("Contributie bank sheet naam is verplicht.")
    elif contributie_bank_path:
        errors.extend(validate_workbook_tabs(contributie_bank_path, [contributie_bank_sheet]))

    tag_targets = contributie.get("tag_targets", {})
    for code in ("8000", "8001"):
        if code not in tag_targets:
            errors.append(f"Contributie bedrag voor {code} ontbreekt.")
            continue
        try:
            float(tag_targets.get(code))
        except (TypeError, ValueError):
            errors.append(f"Contributie bedrag voor {code} is ongeldig.")

    return errors

APPS = {
    "bankrekening": {
        "id": "bankrekening",
        "name": "Bankrekening transacties",
        "description": "Beheer bank- en spaarrekening",
        "cwd": os.path.join(APP_DIR, "project-debutade-bankrekening - v2"),
        "script": "webapp.py",
        "python": os.path.join(".venv", "Scripts", "python.exe"),
        "port": SUBAPP_PORT,
    },
    "kasboek": {
        "id": "kasboek",
        "name": "Kasboek beheer",
        "description": "Bijhouden van Debutade kas",
        "cwd": os.path.join(APP_DIR, "project-debutade-kasboek"),
        "script": "webapp.py",
        "python": os.path.join(".venv", "Scripts", "python.exe"),
        "port": SUBAPP_PORT,
    },
    "bontoevoegen": {
        "id": "bontoevoegen",
        "name": "Bon toevoegen",
        "description": "Voeg bonnen en URLs toe aan transacties",
        "cwd": os.path.join(APP_DIR, "project-debutade-bontoevoegen"),
        "script": "voegbontoe_webapp.py",
        "python": os.path.join(".venv", "Scripts", "python.exe"),
        "port": SUBAPP_PORT,
    },
    "showreport": {
        "id": "showreport",
        "name": "Show rapport",
        "description": "Power BI rapportage",
        "cwd": os.path.join(APP_DIR, "project-debutade-showreport"),
        "script": "webapp.py",
        "python": os.path.join(".venv", "Scripts", "python.exe"),
        "port": SUBAPP_PORT,
    },
    "contributie": {
        "id": "contributie",
        "name": "Contributie overzicht",
        "description": "Koppelt contributies aan leden",
        "cwd": os.path.join(APP_DIR, "project-debutade-contributie"),
        "script": "webapp.py",
        "python": os.path.join(".venv", "Scripts", "python.exe"),
        "port": SUBAPP_PORT,
    },
}

RUNNING_PROCS = {}

app = Flask(
    __name__,
    template_folder=os.path.join(APP_DIR, "templates"),
    static_folder=os.path.join(APP_DIR, "static"),
)


def is_port_open(port, host="127.0.0.1", timeout=0.4):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(timeout)
        return sock.connect_ex((host, port)) == 0


def wait_for_port(port, timeout=SUBAPP_START_TIMEOUT, proc=None):
    start = time.time()
    while time.time() - start < timeout:
        if is_port_open(port):
            return True
        if proc is not None and proc.poll() is not None:
            return False
        time.sleep(0.4)
    return False


def resolve_python_path(app_info):
    candidates = []
    python_rel = app_info.get("python")
    if python_rel:
        candidates.append(os.path.join(app_info["cwd"], python_rel))

    for env_dir in (".venv", "venv", ".venv312"):
        candidates.append(os.path.join(app_info["cwd"], env_dir, "Scripts", "python.exe"))

    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate

    return sys.executable


def ensure_app_running(app_id):
    app_info = APPS.get(app_id)
    if not app_info:
        return None, "Onbekende app"

    if is_port_open(app_info["port"], host="127.0.0.1"):
        if find_app_processes(app_info) or (
            RUNNING_PROCS.get(app_id) and RUNNING_PROCS[app_id].poll() is None
        ):
            return f"http://127.0.0.1:{app_info['port']}", None
        stop_other_apps(app_id)
        time.sleep(0.6)
        if is_port_open(app_info["port"], host="127.0.0.1"):
            return None, f"Poort {app_info['port']} is al in gebruik door een ander programma"

    script_path = os.path.join(app_info["cwd"], app_info["script"])
    if not os.path.exists(script_path):
        return None, f"Script niet gevonden: {script_path}"

    env = os.environ.copy()
    env["MAIN_APP_URL"] = MAIN_APP_URL
    env["DEBUTADE_CONFIG"] = CONFIG_JSON_PATH
    env["DEBUTADE_APP_PORT"] = str(app_info["port"])

    python_path = resolve_python_path(app_info)
    if not os.path.exists(python_path):
        logging.warning("Python niet gevonden (%s), fallback naar %s", python_path, sys.executable)
        python_path = sys.executable

    creationflags = 0
    startupinfo = None
    if os.name == "nt":
        creationflags = subprocess.CREATE_NO_WINDOW
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW

    try:
        proc = subprocess.Popen(
            [python_path, script_path],
            cwd=app_info["cwd"],
            env=env,
            creationflags=creationflags,
            startupinfo=startupinfo,
        )
        RUNNING_PROCS[app_id] = proc
        logging.info("%s gestart (PID: %s)", app_info["name"], proc.pid)
    except Exception as exc:
        logging.exception("Kon %s niet starten", app_info["name"])
        return None, f"Starten mislukt: {exc}"

    if not wait_for_port(app_info["port"], proc=proc):
        if proc.poll() is not None:
            return None, f"App stopte direct (exit code {proc.returncode})"
        return None, "App startte niet binnen de timeout"

    return f"http://127.0.0.1:{app_info['port']}", None


def find_app_processes(app_info):
    if not psutil:
        return []

    procs = []
    target_script = app_info["script"].lower()
    target_dir = os.path.abspath(app_info["cwd"]).lower()

    for proc in psutil.process_iter(["pid", "name", "cmdline", "cwd"]):
        try:
            name = proc.info.get("name") or ""
            if "python" not in name.lower():
                continue
            cmdline = proc.info.get("cmdline") or []
            if target_script not in " ".join(cmdline).lower():
                continue
            cwd = proc.info.get("cwd") or ""
            if cwd and target_dir not in cwd.lower():
                continue
            procs.append(proc)
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            continue

    return procs


def stop_app_processes(app_id):
    app_info = APPS.get(app_id)
    if not app_info:
        return 0

    stopped = 0

    proc = RUNNING_PROCS.get(app_id)
    if proc and proc.poll() is None:
        try:
            proc.terminate()
            proc.wait(timeout=4)
            stopped += 1
        except Exception:
            try:
                proc.kill()
                proc.wait(timeout=2)
                stopped += 1
            except Exception:
                pass

    for other_proc in find_app_processes(app_info):
        try:
            other_proc.terminate()
            other_proc.wait(timeout=4)
            stopped += 1
        except Exception:
            try:
                other_proc.kill()
                other_proc.wait(timeout=2)
                stopped += 1
            except Exception:
                pass

    if app_id in RUNNING_PROCS and RUNNING_PROCS[app_id].poll() is not None:
        RUNNING_PROCS.pop(app_id, None)

    return stopped


def stop_other_apps(current_id):
    stopped = 0
    for other_id in APPS:
        if other_id == current_id:
            continue
        stopped += stop_app_processes(other_id)
    return stopped


def get_app_status(app_id, app_info):
    pid = None

    proc = RUNNING_PROCS.get(app_id)
    if proc and proc.poll() is None:
        pid = proc.pid
    elif proc:
        RUNNING_PROCS.pop(app_id, None)

    if pid is None:
        procs = find_app_processes(app_info)
        if procs:
            pid = procs[0].pid

    return {
        "running": pid is not None,
        "port": app_info["port"],
        "pid": pid,
    }


@app.route("/")
def index():
    apps = [APPS[key] for key in APPS]
    error = request.args.get("error")
    return render_template(
        "main.html",
        apps=apps,
        error=error,
        main_url=MAIN_APP_URL,
        today=datetime.now().strftime("%d-%m-%Y"),
    )


@app.route("/help")
def help_main():
    return render_template("help_main.html")


@app.route("/settings", methods=["GET", "POST"])
def settings_main():
    saved = request.args.get("saved") == "1"
    errors = []
    config_data, load_error = load_main_config()

    if request.method == "POST":
        if load_error:
            return render_template(
                "settings_main.html",
                config_data={},
                errors=[load_error],
                saved=False,
            )

        def form_value(key):
            return (request.form.get(key) or "").strip()

        existing_config = config_data or {}
        existing_shared = existing_config.get("shared", {})
        existing_bank = existing_config.get("bankrekening", {})
        existing_kas = existing_config.get("kasboek", {})
        existing_bon = existing_config.get("bontoevoegen", {})
        existing_showreport = existing_config.get("showreport", {})
        existing_contributie = existing_config.get("contributie", {})

        shared_bank_excel = form_value("shared_bank_excel_file_name") or existing_shared.get("bank_excel_file_name")

        shared_config = {
            "grootboek_directory": form_value("shared_grootboek_directory"),
            "bank_excel_file_name": shared_bank_excel or "",
            "backup_directory": form_value("shared_backup_directory"),
            "log_directory": form_value("shared_log_directory"),
            "resources": form_value("shared_resources"),
            "log_level": form_value("shared_log_level").upper() or "INFO",
            "tags": split_lines(form_value("shared_tags")),
        }

        bank_config = {
            "excel_file_name": form_value("bank_excel_file_name") or shared_bank_excel or existing_bank.get("excel_file_name"),
            "excel_sheet_name": form_value("bank_excel_sheet_name"),
            "required_sheets": split_lines(form_value("bank_required_sheets")),
        }

        kas_config = {
            "excel_file_name": form_value("kas_excel_file_name") or existing_kas.get("excel_file_name"),
            "excel_sheet_name": form_value("kas_excel_sheet_name") or existing_kas.get("excel_sheet_name"),
        }

        bon_config = {
            "bank_excel_file_name": form_value("bon_bank_excel_file_name") or shared_bank_excel or existing_bon.get("bank_excel_file_name"),
            "kas_excel_file_name": form_value("bon_kas_excel_file_name") or existing_bon.get("kas_excel_file_name"),
            "sharepoint_tenant": form_value("bon_sharepoint_tenant") or existing_bon.get("sharepoint_tenant", ""),
        }

        showreport_config = {
            "report_url": form_value("showreport_report_url") or existing_showreport.get("report_url", ""),
            "report_title": form_value("showreport_report_title") or existing_showreport.get("report_title", ""),
        }
        contributie_config = dict(existing_contributie)
        contributie_config.update({
            "ledenbestand_path": form_value("contrib_ledenbestand_path") or existing_contributie.get("ledenbestand_path", ""),
            "leden_sheet_personen": form_value("contrib_leden_sheet_personen") or existing_contributie.get("leden_sheet_personen", ""),
            "leden_sheet_betaald": form_value("contrib_leden_sheet_betaald") or existing_contributie.get("leden_sheet_betaald", ""),
            "bank_excel_file_name": form_value("contrib_bank_excel_file_name") or shared_bank_excel or existing_contributie.get("bank_excel_file_name"),
            "bank_sheet_name": form_value("contrib_bank_sheet_name") or existing_contributie.get("bank_sheet_name", ""),
            "tags": split_lines(form_value("contrib_tags")) or existing_contributie.get("tags", []),
            "tag_targets": {
                "8000": form_value("contrib_target_8000") or existing_contributie.get("tag_targets", {}).get("8000", ""),
                "8001": form_value("contrib_target_8001") or existing_contributie.get("tag_targets", {}).get("8001", ""),
            },
        })

        new_config = {
            "shared": shared_config,
            "bankrekening": bank_config,
            "kasboek": kas_config,
            "bontoevoegen": bon_config,
            "showreport": showreport_config,
            "contributie": contributie_config,
        }

        errors = validate_main_config(new_config)
        if errors:
            config_data = new_config
        else:
            save_error = save_main_config(new_config)
            if save_error:
                errors = [save_error]
                config_data = new_config
            else:
                return redirect(url_for("settings_main", saved="1"))

    if load_error:
        errors = [load_error]
        config_data = {}
    elif config_data:
        errors = validate_main_config(config_data)

    return render_template(
        "settings_main.html",
        config_data=config_data or {},
        errors=errors,
        saved=saved,
    )


@app.route("/launch/<app_id>")
def launch(app_id):
    url, error = ensure_app_running(app_id)
    if error:
        return redirect(url_for("index", error=error))
    return redirect(url)


@app.route("/status")
def status():
    return jsonify({
        app_id: get_app_status(app_id, app_info)
        for app_id, app_info in APPS.items()
    })


@app.route("/stop/<app_id>", methods=["POST"])
def stop(app_id):
    stopped = stop_app_processes(app_id)
    if not stopped:
        return jsonify({"success": False, "message": "Geen proces gevonden"}), 404
    return jsonify({"success": True, "stopped": stopped})


@app.route("/quit", methods=["POST"])
def quit_app():
    return jsonify({"success": True, "message": "Hoofdapp blijft actief"}), 200


if __name__ == "__main__":
    logging.info("Debutade hoofdapp gestart op %s", MAIN_APP_URL)
    app.run(debug=False, host=MAIN_HOST, port=MAIN_PORT)