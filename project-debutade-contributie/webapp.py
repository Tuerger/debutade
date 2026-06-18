"""
Contributie Debutade - Web Applicatie
====================================

Web applicatie voor het koppelen van contributiebetalingen aan leden.

Functionaliteiten:
- Leest ledengegevens uit Ledenbestand.xlsx (tab leden)
- Leest banktransacties uit tab bankrekening en zoekt ID-lid in mededelingen
- Maakt een overzicht met Te innen bedrag, Ontvangen bedrag en status

Versie: 1.0
Datum: 2026-02-16
Auteur: Eric G.
"""

from flask import Flask, render_template, jsonify, request, g
from openpyxl import load_workbook
from datetime import datetime
import logging
import os
import json
import sys
import re
import tempfile
import shutil
import time

# Fix encoding voor Windows console
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

app = Flask(__name__)
app.config["TEMPLATES_AUTO_RELOAD"] = True

CACHE_TTL_SECONDS = int(os.getenv("DEBUTADE_CACHE_TTL_SECONDS", "20"))
BENCHMARK_ENABLED = os.getenv("DEBUTADE_BENCHMARK", "1") == "1"
RUNTIME_CACHE = {}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.getenv(
    "DEBUTADE_CONFIG",
    os.path.abspath(os.path.join(SCRIPT_DIR, "..", "config.json")),
)


def load_config(config_path, section_key="contributie"):
    """Laad de configuratie uit config.json."""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Configuratiebestand niet gevonden: {config_path}")

    with open(config_path, "r", encoding="utf-8") as config_file:
        root_config = json.load(config_file)

    if section_key not in root_config:
        raise KeyError(f"Configuratiesectie ontbreekt: {section_key}")

    config = root_config[section_key]
    shared = root_config.get("shared", {})
    for key in ("backup_directory", "log_directory", "log_level"):
        if key in shared:
            config[key] = shared[key]
    if shared.get("bank_excel_file_name") and not config.get("bank_excel_file_name"):
        config["bank_excel_file_name"] = shared["bank_excel_file_name"]

    if shared.get("grootboek_directory") and config.get("bank_excel_file_name"):
        config["bank_excel_file_path"] = os.path.join(
            shared["grootboek_directory"],
            config["bank_excel_file_name"],
        )

    required_keys = [
        "ledenbestand_path",
        "bank_excel_file_name",
        "bank_sheet_name",
    ]
    for key in required_keys:
        if key not in config:
            raise KeyError(f"Configuratiesleutel ontbreekt: {key}")

    return config


try:
    config = load_config(CONFIG_PATH)
except (FileNotFoundError, KeyError) as exc:
    print(f"WAARSCHUWING: {exc}")
    config = {
        "ledenbestand_path": r"C:\pad\naar\Ledenbestand.xlsx",
        "leden_sheet_name": "leden",
        "bank_excel_file_name": "Debutade boekjaar 2026 Bank.xlsx",
        "bank_sheet_name": "bankrekening",
        "backup_directory": os.path.join(SCRIPT_DIR, "backup"),
        "log_directory": os.path.join(SCRIPT_DIR, "logs"),
        "log_level": "INFO",
    }

LEDENBESTAND_PATH = config["ledenbestand_path"]
LEDEN_SHEET_NAME = (
    config.get("leden_sheet_name")
    or config.get("leden_sheet_leden")
    or "leden"
)
BANK_EXCEL_PATH = config.get("bank_excel_file_path") or config.get("bank_excel_file_name")
BANK_SHEET_NAME = config["bank_sheet_name"]
MANUAL_TRANSACTION_MAPPINGS = config.get("manual_transaction_mappings", {})
MANUAL_PAID_OVERRIDES = config.get("manual_paid_overrides", {})
MANUAL_SPLIT_DECISIONS = config.get("manual_split_decisions", {})
BANK_EXCEL_FALLBACK_BASENAMES = [
    "Debutade boekjaar 2026 Bank",
    "Debutade boekjaar bank 2026",
]
SHARED_BANK_EXCEL_FILE_NAME = config.get("bank_excel_file_name", "")
ALLOWED_CONTRIBUTIE_TAG_CODES = {"8000", "8001"}
ALLOWED_REFUND_TAG_CODES = {"4990", "4991"}
MANUAL_REFUND_OVERRIDES = config.get("manual_refund_overrides", {})
PROCESSED_ORPHANED_REFUNDS = config.get("processed_orphaned_refunds", {})
BACKUP_DIRECTORY = config.get("backup_directory", os.path.join(SCRIPT_DIR, "backup"))
LOG_DIRECTORY = config.get("log_directory", os.path.join(SCRIPT_DIR, "logs"))
LOG_LEVEL = config.get("log_level", "INFO")
MAIN_APP_URL = os.getenv("MAIN_APP_URL", "").strip()

if not os.path.exists(LOG_DIRECTORY):
    os.makedirs(LOG_DIRECTORY)

log_file = os.path.join(LOG_DIRECTORY, f"contributie_{datetime.now().strftime('%Y%m%d')}.log")
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL),
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)


def normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "").replace("-", "")


def _file_signature(file_path):
    if not file_path or not os.path.exists(file_path):
        return None
    stat = os.stat(file_path)
    return (stat.st_mtime_ns, stat.st_size)


def _cache_get(key):
    entry = RUNTIME_CACHE.get(key)
    if not entry:
        return None

    if time.time() - entry["ts"] > CACHE_TTL_SECONDS:
        RUNTIME_CACHE.pop(key, None)
        return None

    return entry["value"]


def _cache_set(key, value):
    RUNTIME_CACHE[key] = {"ts": time.time(), "value": value}
    return value


def invalidate_runtime_cache():
    RUNTIME_CACHE.clear()


@app.before_request
def _benchmark_start():
    if BENCHMARK_ENABLED:
        g._start_time = time.perf_counter()


@app.after_request
def _benchmark_end(response):
    if BENCHMARK_ENABLED and hasattr(g, "_start_time"):
        elapsed_ms = (time.perf_counter() - g._start_time) * 1000
        logging.info("PERF %s %s %s %.1fms", request.method, request.path, response.status_code, elapsed_ms)
    return response


def normalize_account(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip().replace(" ", "")


def normalize_lastname(value):
    return str(value or "").strip().lower()


def build_refund_key(mededelingen, amount):
    """Build a stable key for orphaned refund persistence."""
    normalized_note = str(mededelingen or "").strip()
    normalized_amount = abs(parse_amount(amount))
    return f"{normalized_note}|{normalized_amount:.2f}"


def extract_iban_from_text(value):
    """Extract and normalize first IBAN from free text or return empty string."""
    text = str(value or "")
    if not text:
        return ""

    upper_text = text.upper()
    # Match IBAN even when surrounded by punctuation or followed by words like "Kenmerk".
    match = re.search(r"NL\s*\d{2}\s*[A-Z]{4}(?:\s*\d){10}", upper_text)
    if not match:
        return ""
    return normalize_account(re.sub(r"\s+", "", match.group(0)))


def extract_name_from_text(value):
    """Extract payer name from free text fields like 'Naam: ... Omschrijving: ...'."""
    text = str(value or "")
    if not text:
        return ""

    name_match = re.search(r"Naam\s*:\s*(.+?)(?:\s+Omschrijving\s*:|\s+IBAN\s*:|$)", text, flags=re.IGNORECASE)
    if not name_match:
        return ""
    return name_match.group(1).strip()


def normalize_name_for_match(value):
    text = str(value or "").lower().strip()
    if not text:
        return ""
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_amount(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    # Verwijder valutatekens en niet-numerieke tekens behalve . , -
    cleaned = re.sub(r"[^0-9,\.\-]", "", text)
    if not cleaned:
        return 0.0

    # Als er een komma en punt zijn, ga uit van punt als duizendtallen
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "")

    # Gebruik punt als decimaal scheiding
    cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def normalize_member_id_4digit(value):
    text = str(value or "").strip()
    if not text:
        return ""

    if re.fullmatch(r"\d+(\.0+)?", text):
        try:
            text = str(int(float(text)))
        except ValueError:
            pass

    exact_match = re.search(r"(?<!\d)(\d{4})(?!\d)", text)
    if exact_match:
        return exact_match.group(1)

    digits_only = "".join(re.findall(r"\d", text))
    if not digits_only:
        return ""
    if len(digits_only) >= 4:
        return digits_only[-4:]
    return digits_only.zfill(4)


def extract_4digit_tokens(value):
    text = str(value or "")
    if not text:
        return set()
    return set(re.findall(r"(?<!\d)\d{4}(?!\d)", text))


def find_manual_mapping_for_transaction(mededelingen):
    if not mededelingen:
        return None

    mededelingen_lower = str(mededelingen).lower()
    for member_id, mapping_fragment in MANUAL_TRANSACTION_MAPPINGS.items():
        fragment_lower = str(mapping_fragment or "").lower()
        if fragment_lower and fragment_lower in mededelingen_lower:
            return member_id

    return None


def find_manual_split_decision(mededelingen):
    if not mededelingen:
        return None

    mededelingen_lower = str(mededelingen).lower()
    best_match = None
    best_match_length = -1

    for fragment, decision in MANUAL_SPLIT_DECISIONS.items():
        fragment_text = str(fragment or "").strip()
        if not fragment_text:
            continue

        fragment_lower = fragment_text.lower()
        if fragment_lower not in mededelingen_lower:
            continue

        if len(fragment_lower) <= best_match_length:
            continue

        if isinstance(decision, str):
            normalized = {"mode": decision}
        elif isinstance(decision, dict):
            normalized = dict(decision)
        else:
            continue

        mode = str(normalized.get("mode", "")).strip().lower()
        if mode not in {"split", "single_member"}:
            continue

        if mode == "single_member" and not str(normalized.get("member_id", "")).strip():
            continue

        best_match = normalized
        best_match_length = len(fragment_lower)

    return best_match


def extract_tag_code(value):
    text = str(value or "").strip().lower()
    if not text:
        return ""
    if ";" in text:
        text = text.split(";", 1)[0]
    return text.strip()


def build_header_map(sheet):
    headers = {}
    for idx, cell in enumerate(sheet[1]):
        normalized = normalize_header(cell.value)
        if normalized:
            headers[normalized] = idx
    return headers


def get_cell_value(row, header_map, *names):
    for name in names:
        idx = header_map.get(normalize_header(name))
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def read_sheet_rows(file_path, sheet_name):
    if not os.path.exists(file_path):
        return [], f"Excel bestand niet gevonden: {file_path}"

    wb = None
    temp_path = None
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except PermissionError:
        # Bestand is waarschijnlijk open in Excel of door OneDrive gelocked.
        try:
            temp_dir = os.path.join(SCRIPT_DIR, "cache")
            os.makedirs(temp_dir, exist_ok=True)
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=temp_dir)
            temp_file.close()
            temp_path = temp_file.name
            shutil.copy2(file_path, temp_path)
            wb = load_workbook(temp_path, read_only=True, data_only=True)
        except Exception:
            return [], f"Bestand is in gebruik: {file_path}. Sluit Excel en probeer opnieuw."
    except Exception as exc:  # noqa: BLE001
        return [], f"Kan Excel bestand niet openen: {file_path} ({exc})"

    matched_sheet_name = None
    for current_name in wb.sheetnames:
        if current_name.strip().lower() == str(sheet_name).strip().lower():
            matched_sheet_name = current_name
            break

    if not matched_sheet_name:
        wb.close()
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)
        return [], f"Tabblad niet gevonden: {sheet_name}"

    sheet = wb[matched_sheet_name]
    header_map = build_header_map(sheet)
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    wb.close()
    if temp_path and os.path.exists(temp_path):
        os.remove(temp_path)
    return (rows, header_map), None


def load_ledenbestand():
    records = []
    result, error = read_sheet_rows(LEDENBESTAND_PATH, LEDEN_SHEET_NAME)
    if error:
        return [], [error]

    rows, header_map = result
    for row in rows:
        member_id = get_cell_value(row, header_map, "ID-lid", "ID lid", "ID")
        achternaam = get_cell_value(row, header_map, "Achternaam")
        email = get_cell_value(row, header_map, "Email", "E-mail", "Mail")
        rekeningnummer = get_cell_value(
            row,
            header_map,
            "Rekeningnummer",
            "Rekening nummer",
            "IBAN",
            "Bankrekening",
        )
        te_innen_bedrag = get_cell_value(
            row,
            header_map,
            "bedrag",
            "Te innen bedrag",
            "Contributie",
        )

        member_id = str(member_id).strip() if member_id is not None else ""
        achternaam = str(achternaam or "").strip()
        email = str(email or "").strip()
        rekeningnummer = normalize_account(rekeningnummer)
        due_amount = parse_amount(te_innen_bedrag)

        if not member_id and not achternaam and not email and due_amount == 0:
            continue

        if not member_id:
            continue

        records.append(
            {
                "member_id": member_id,
                "member_id_4digit": normalize_member_id_4digit(member_id),
                "achternaam": achternaam,
                "email": email,
                "rekeningnummer": rekeningnummer,
                "due_amount": due_amount,
                "received_amount": 0.0,
                "matched_transactions": [],
                "opmerking": "",
                "status_icon": "❌",
                "status_label": "Nog niets ontvangen",
                "status_class": "status-none",
            }
        )

    return records, []


def load_bank_transactions():
    bank_excel_path = resolve_bank_excel_path()
    result, error = read_sheet_rows(bank_excel_path, BANK_SHEET_NAME)
    if error:
        return [], [error]

    rows, header_map = result
    transactions = []
    for row in rows:
        tag_value = get_cell_value(row, header_map, "Tag", "Tags")
        tag_code = extract_tag_code(tag_value)
        if tag_code not in ALLOWED_CONTRIBUTIE_TAG_CODES:
            continue

        mededelingen = get_cell_value(
            row,
            header_map,
            "Mededelingen",
            "Omschrijving",
            "Beschrijving",
        )
        tegenrekening = get_cell_value(
            row,
            header_map,
            "Tegenrekening",
            "Tegen rekening",
            "Rekening",
            "IBAN",
        )
        bedrag = get_cell_value(row, header_map, "Bedrag", "Bedrag (EUR)", "BedragEUR")
        af_bij = get_cell_value(row, header_map, "Af Bij", "Af/Bij")

        amount_value = parse_amount(bedrag)
        if str(af_bij or "").strip().lower() == "af":
            amount_value = -amount_value

        if not mededelingen and amount_value == 0:
            continue

        transactions.append(
            {
                "mededelingen": str(mededelingen or ""),
                "amount": amount_value,
                "tegenrekening": normalize_account(tegenrekening),
            }
        )

    return transactions, []


def load_refund_transactions():
    """Laad terugstortingstransacties (4990, 4991)."""
    bank_excel_path = resolve_bank_excel_path()
    result, error = read_sheet_rows(bank_excel_path, BANK_SHEET_NAME)
    if error:
        return [], [error]

    rows, header_map = result
    refunds = []
    for row in rows:
        tag_value = get_cell_value(row, header_map, "Tag", "Tags")
        tag_code = extract_tag_code(tag_value)
        if tag_code not in ALLOWED_REFUND_TAG_CODES:
            continue

        mededelingen = get_cell_value(
            row,
            header_map,
            "Mededelingen",
            "Omschrijving",
            "Beschrijving",
        )
        tegenrekening = get_cell_value(
            row,
            header_map,
            "Tegenrekening",
            "Tegen rekening",
            "Rekening",
            "IBAN",
        )
        bedrag = get_cell_value(row, header_map, "Bedrag", "Bedrag (EUR)", "BedragEUR")
        af_bij = get_cell_value(row, header_map, "Af Bij", "Af/Bij")

        amount_value = parse_amount(bedrag)
        if str(af_bij or "").strip().lower() == "af":
            amount_value = -amount_value

        if not mededelingen and amount_value == 0:
            continue

        refunds.append(
            {
                "mededelingen": str(mededelingen or ""),
                "amount": amount_value,
                "tegenrekening": normalize_account(tegenrekening),
            }
        )

    return refunds, []


def build_transaction_totals_by_member_id_4digit(transactions):
    totals_by_member_id_4digit = {}

    for transaction in transactions:
        amount = transaction.get("amount", 0.0)
        mededelingen = transaction.get("mededelingen", "")
        id_tokens = extract_4digit_tokens(mededelingen)

        for token in id_tokens:
            totals_by_member_id_4digit[token] = totals_by_member_id_4digit.get(token, 0.0) + amount

    return totals_by_member_id_4digit


def split_amount_equally(amount, parts):
    if parts <= 0:
        return []

    amount_cents = int(round(float(amount) * 100))
    sign = -1 if amount_cents < 0 else 1
    abs_cents = abs(amount_cents)
    base = abs_cents // parts
    remainder = abs_cents % parts

    shares = []
    for idx in range(parts):
        cents = (base + (1 if idx < remainder else 0)) * sign
        shares.append(cents / 100.0)
    return shares


def build_id_split_allocations(transactions, member_lookup_by_4digit):
    allocations_by_member = {}
    transaction_has_known_member_id = {}
    pending_ambiguous_transactions = []

    for idx, transaction in enumerate(transactions):
        mededelingen = transaction.get("mededelingen", "")
        tokens = extract_4digit_tokens(mededelingen)
        matched_tokens = sorted(token for token in tokens if token in member_lookup_by_4digit)

        if not matched_tokens:
            transaction_has_known_member_id[idx] = False
            continue

        transaction_has_known_member_id[idx] = True

        if len(matched_tokens) == 1:
            token = matched_tokens[0]
            allocations_by_member.setdefault(token, []).append(
                {
                    "index": idx,
                    "amount": round(float(transaction.get("amount", 0.0)), 2),
                    "split_count": 1,
                }
            )
            continue

        decision = find_manual_split_decision(mededelingen)
        if decision and decision.get("mode") == "split":
            shares = split_amount_equally(transaction.get("amount", 0.0), len(matched_tokens))
            for token, share in zip(matched_tokens, shares):
                allocations_by_member.setdefault(token, []).append(
                    {
                        "index": idx,
                        "amount": round(float(share), 2),
                        "split_count": len(matched_tokens),
                    }
                )
            continue

        if decision and decision.get("mode") == "single_member":
            selected_member_id = str(decision.get("member_id", "")).strip()
            selected_member_id_4digit = normalize_member_id_4digit(selected_member_id)
            if selected_member_id_4digit in matched_tokens:
                allocations_by_member.setdefault(selected_member_id_4digit, []).append(
                    {
                        "index": idx,
                        "amount": round(float(transaction.get("amount", 0.0)), 2),
                        "split_count": 1,
                    }
                )
                continue

        pending_ambiguous_transactions.append(
            {
                "index": idx,
                "amount": round(float(transaction.get("amount", 0.0)), 2),
                "mededelingen": str(mededelingen or ""),
                "candidates": [
                    {
                        "member_id": str(member_lookup_by_4digit[token].get("member_id", "")),
                        "achternaam": str(member_lookup_by_4digit[token].get("achternaam", "")),
                    }
                    for token in matched_tokens
                ],
            }
        )

    return allocations_by_member, transaction_has_known_member_id, pending_ambiguous_transactions


def build_name_pattern(name_value):
    achternaam = str(name_value or "").strip()
    if not achternaam:
        return None

    if len(achternaam) < 3:
        return None

    escaped = re.escape(achternaam)
    escaped = escaped.replace(r"\ ", r"\s+")
    pattern = rf"(?<![A-Za-zÀ-ÿ]){escaped}(?![A-Za-zÀ-ÿ])"
    return re.compile(pattern, flags=re.IGNORECASE)


def calculate_received_by_name_fallback(achternaam, transactions):
    pattern = build_name_pattern(achternaam)
    if pattern is None:
        return 0.0, False, 0

    fallback_amount = 0.0
    found = False
    matched_count = 0
    for transaction in transactions:
        mededelingen = str(transaction.get("mededelingen", ""))
        if not mededelingen:
            continue

        if pattern.search(mededelingen):
            fallback_amount += transaction.get("amount", 0.0)
            found = True
            matched_count += 1

    return round(fallback_amount, 2), found, matched_count


def collect_matched_transaction_indexes(record, transactions):
    """Collect matching transaction indexes for a member."""
    member_id = str(record.get("member_id", "")).strip()
    member_id_4digit = str(record.get("member_id_4digit", "")).strip()
    achternaam = str(record.get("achternaam", "")).strip()

    matched_by_id = set()
    matched_by_manual = set()
    matched_by_fallback = set()

    for i, tx in enumerate(transactions):
        mededelingen = str(tx.get("mededelingen", ""))

        if member_id_4digit and member_id_4digit in extract_4digit_tokens(mededelingen):
            matched_by_id.add(i)

        if find_manual_mapping_for_transaction(mededelingen) == member_id:
            matched_by_manual.add(i)

    all_matches = set()
    all_matches.update(matched_by_id)
    all_matches.update(matched_by_manual)

    # Use name fallback only when no ID/manual matches are found.
    if not all_matches:
        pattern = build_name_pattern(achternaam)
        if pattern:
            for i, tx in enumerate(transactions):
                mededelingen = str(tx.get("mededelingen", ""))
                if mededelingen and pattern.search(mededelingen):
                    matched_by_fallback.add(i)
            all_matches.update(matched_by_fallback)

    return {
        "all": all_matches,
        "id": matched_by_id,
        "manual": matched_by_manual,
        "fallback": matched_by_fallback,
    }


def resolve_bank_excel_path():
    if BANK_EXCEL_PATH and os.path.exists(BANK_EXCEL_PATH):
        return BANK_EXCEL_PATH

    if SHARED_BANK_EXCEL_FILE_NAME and BANK_EXCEL_PATH:
        bank_dir = os.path.dirname(BANK_EXCEL_PATH)
        shared_candidate = os.path.join(bank_dir, SHARED_BANK_EXCEL_FILE_NAME)
        if os.path.exists(shared_candidate):
            return shared_candidate

    search_dirs = []
    if BANK_EXCEL_PATH:
        bank_dir = os.path.dirname(BANK_EXCEL_PATH)
        if bank_dir:
            search_dirs.append(bank_dir)

    if SCRIPT_DIR not in search_dirs:
        search_dirs.append(SCRIPT_DIR)

    for search_dir in search_dirs:
        for fallback_basename in BANK_EXCEL_FALLBACK_BASENAMES:
            for extension in (".xlsx", ".xlsm", ".xls"):
                candidate = os.path.join(search_dir, f"{fallback_basename}{extension}")
                if os.path.exists(candidate):
                    return candidate

        try:
            for filename in os.listdir(search_dir):
                filename_lower = filename.lower()
                if "debutade" in filename_lower and "boekjaar" in filename_lower and "bank" in filename_lower and "2026" in filename_lower:
                    wildcard_candidate = os.path.join(search_dir, filename)
                    if os.path.isfile(wildcard_candidate):
                        return wildcard_candidate
        except OSError:
            continue

    return BANK_EXCEL_PATH


def load_processed_orphaned_refunds():
    """Normalize processed orphaned refunds from config for matching/reporting."""
    processed = []
    for refund_key, payload in PROCESSED_ORPHANED_REFUNDS.items():
        if not isinstance(payload, dict):
            continue

        amount = round(abs(parse_amount(payload.get("amount", 0.0))), 2)
        if amount <= 0:
            continue

        reason = str(payload.get("reason", "")).strip()
        processed_at = str(payload.get("processed_at", "")).strip()
        mededelingen = str(refund_key).split("|", 1)[0] if "|" in str(refund_key) else str(refund_key)
        refund_iban = extract_iban_from_text(mededelingen)

        processed.append(
            {
                "refund_key": str(refund_key),
                "mededelingen": mededelingen,
                "amount": amount,
                "reason": reason,
                "processed_at": processed_at,
                "iban": refund_iban,
                "penningmeester_opmerking": f"€ {amount:.2f} - {reason}" if reason else f"€ {amount:.2f}",
            }
        )

    return processed


def build_overview():
    bank_excel_path = resolve_bank_excel_path()
    cache_key = (
        "contributie_overview",
        _file_signature(LEDENBESTAND_PATH),
        _file_signature(bank_excel_path),
        _file_signature(CONFIG_PATH),
    )
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    records, errors = load_ledenbestand()
    transactions, bank_errors = load_bank_transactions()
    refunds, refund_errors = load_refund_transactions()
    errors.extend(bank_errors)
    errors.extend(refund_errors)

    lastname_counts = {}
    for record in records:
        lastname_key = normalize_lastname(record.get("achternaam", ""))
        if lastname_key:
            lastname_counts[lastname_key] = lastname_counts.get(lastname_key, 0) + 1

    fallback_assigned_payment_indexes = set()
    fallback_assigned_refund_indexes = set()
    processed_orphaned_entries = load_processed_orphaned_refunds()
    processed_with_meta = [
        {
            "refund": item,
            "iban": normalize_account(item.get("iban", "")) or extract_iban_from_text(item.get("mededelingen", "")),
            "name": normalize_name_for_match(extract_name_from_text(item.get("mededelingen", ""))),
            "used": False,
        }
        for item in processed_orphaned_entries
    ]

    # Build a lookup for refunds by 4-digit ID
    refunds_by_member_id_4digit = {}
    for idx, refund in enumerate(refunds):
        mededelingen = str(refund.get("mededelingen", ""))
        id_tokens = extract_4digit_tokens(mededelingen)
        for token in id_tokens:
            refunds_by_member_id_4digit.setdefault(token, []).append((idx, refund))

    member_lookup_by_4digit = {
        str(record.get("member_id_4digit", "")).strip(): {
            "member_id": str(record.get("member_id", "")).strip(),
            "achternaam": str(record.get("achternaam", "")).strip(),
        }
        for record in records
        if str(record.get("member_id_4digit", "")).strip()
    }
    id_split_allocations, transaction_has_known_member_id, pending_ambiguous_transactions = build_id_split_allocations(
        transactions,
        member_lookup_by_4digit,
    )

    refund_has_known_member_id = {}
    for idx, refund in enumerate(refunds):
        tokens = extract_4digit_tokens(refund.get("mededelingen", ""))
        refund_has_known_member_id[idx] = any(token in member_lookup_by_4digit for token in tokens)

    matched_transaction_keys = set()
    matched_refund_indexes = set()

    exact_count = 0
    partial_count = 0
    none_count = 0
    neutral_count = 0

    for record in records:
        member_id = str(record.get("member_id", "")).strip()
        member_id_4digit = record.get("member_id_4digit", "")
        member_rekeningnummer = normalize_account(record.get("rekeningnummer", ""))
        member_lastname_key = normalize_lastname(record.get("achternaam", ""))
        lastname_is_unique = lastname_counts.get(member_lastname_key, 0) == 1 if member_lastname_key else False
        due_amount = record.get("due_amount", 0.0)
        record["received_amount"] = 0.0
        record["refunded_amount"] = 0.0
        record["matched_transactions"] = []
        record["matched_refunds"] = []
        record["tegenrekening"] = ""
        record["opmerking"] = ""
        record["penningmeester_opmerking"] = ""
        record["manual_paid_override"] = False
        record["manual_paid_reason"] = ""
        record["has_refund_transactions"] = False
        record["is_overledene_opzegger"] = False

        override_entry = MANUAL_PAID_OVERRIDES.get(member_id)
        if isinstance(override_entry, dict) and override_entry.get("marked_paid"):
            reason = str(override_entry.get("reason", "")).strip()
            record["manual_paid_override"] = True
            record["manual_paid_reason"] = reason
            if reason:
                record["penningmeester_opmerking"] = f"Betaald-markering: {reason}"

        if not member_id_4digit:
            record["opmerking"] = "ID-lid bevat geen 4 cijfers"
            record["status_icon"] = "❌"
            record["status_label"] = "Geen geldig 4-cijferig ID"
            record["status_class"] = "status-none"
            none_count += 1
            continue

        matched_by_id = set()
        matched_by_manual = set()
        matched_by_fallback = set()
        matched_entries = {}

        for allocation in id_split_allocations.get(member_id_4digit, []):
            tx_index = allocation["index"]
            matched_by_id.add(tx_index)
            tx = transactions[tx_index]
            matched_entries[tx_index] = {
                "amount": round(float(allocation.get("amount", 0.0)), 2),
                "mededelingen": str(tx.get("mededelingen", "")),
                "tegenrekening": str(tx.get("tegenrekening", "")),
                "split_count": int(allocation.get("split_count", 1)),
            }

        for i, tx in enumerate(transactions):
            if i in matched_by_id:
                continue

            mededelingen = str(tx.get("mededelingen", ""))
            if find_manual_mapping_for_transaction(mededelingen) != member_id:
                continue

            # Manual mapping is fallback: do not override transactions with known IDs.
            if transaction_has_known_member_id.get(i, False):
                continue

            matched_by_manual.add(i)
            matched_entries[i] = {
                "amount": round(float(tx.get("amount", 0.0)), 2),
                "mededelingen": str(tx.get("mededelingen", "")),
                "tegenrekening": str(tx.get("tegenrekening", "")),
                "split_count": 1,
            }

        all_matches = set()
        all_matches.update(matched_by_id)
        all_matches.update(matched_by_manual)

        if not all_matches and lastname_is_unique:
            pattern = build_name_pattern(record.get("achternaam", ""))
            if pattern:
                for i, tx in enumerate(transactions):
                    if transaction_has_known_member_id.get(i, False):
                        continue
                    if i in fallback_assigned_payment_indexes:
                        continue
                    mededelingen = str(tx.get("mededelingen", ""))
                    if mededelingen and pattern.search(mededelingen):
                        matched_by_fallback.add(i)
                        matched_entries[i] = {
                            "amount": round(float(tx.get("amount", 0.0)), 2),
                            "mededelingen": mededelingen,
                            "tegenrekening": str(tx.get("tegenrekening", "")),
                            "split_count": 1,
                        }
                fallback_assigned_payment_indexes.update(matched_by_fallback)
                all_matches.update(matched_by_fallback)

        matched_by_iban = set()
        if member_rekeningnummer:
            for i, tx in enumerate(transactions):
                if i in fallback_assigned_payment_indexes:
                    continue
                if i in all_matches:
                    continue
                if transaction_has_known_member_id.get(i, False):
                    continue
                tx_tegenrekening = normalize_account(tx.get("tegenrekening", ""))
                if tx_tegenrekening and tx_tegenrekening == member_rekeningnummer:
                    matched_by_iban.add(i)
                    matched_entries[i] = {
                        "amount": round(float(tx.get("amount", 0.0)), 2),
                        "mededelingen": str(tx.get("mededelingen", "")),
                        "tegenrekening": str(tx.get("tegenrekening", "")),
                        "split_count": 1,
                    }

            if matched_by_iban:
                fallback_assigned_payment_indexes.update(matched_by_iban)
                all_matches.update(matched_by_iban)

        all_indexes = sorted(all_matches)
        matched_items = [matched_entries[i] for i in all_indexes if i in matched_entries]
        received_amount = round(sum(item.get("amount", 0.0) for item in matched_items), 2)
        record["received_amount"] = received_amount
        record["matched_transactions"] = matched_items

        refund_match_ibans = set()
        if member_rekeningnummer:
            refund_match_ibans.add(member_rekeningnummer)
        for item in matched_items:
            tx_iban = normalize_account(item.get("tegenrekening", ""))
            if tx_iban:
                refund_match_ibans.add(tx_iban)

        # Process refunds
        refund_items = []
        if member_id_4digit and member_id_4digit in refunds_by_member_id_4digit:
            for refund_index, refund in refunds_by_member_id_4digit[member_id_4digit]:
                refund_items.append({
                    "amount": round(abs(parse_amount(refund.get("amount", 0.0))), 2),
                    "mededelingen": str(refund.get("mededelingen", "")),
                    "tegenrekening": str(refund.get("tegenrekening", "")),
                    "index": refund_index,
                })
                matched_refund_indexes.add(refund_index)
                record["has_refund_transactions"] = True

        if refund_match_ibans:
            for idx, refund in enumerate(refunds):
                if idx in fallback_assigned_refund_indexes:
                    continue
                if idx in matched_refund_indexes:
                    continue
                if refund_has_known_member_id.get(idx, False):
                    continue
                tx_tegenrekening = normalize_account(refund.get("tegenrekening", ""))
                if tx_tegenrekening and tx_tegenrekening in refund_match_ibans:
                    refund_items.append({
                        "amount": round(abs(parse_amount(refund.get("amount", 0.0))), 2),
                        "mededelingen": str(refund.get("mededelingen", "")),
                        "tegenrekening": str(refund.get("tegenrekening", "")),
                        "index": idx,
                    })
                    fallback_assigned_refund_indexes.add(idx)
                    matched_refund_indexes.add(idx)
                    record["has_refund_transactions"] = True

        if not refund_items and lastname_is_unique:
            pattern = build_name_pattern(record.get("achternaam", ""))
            if pattern:
                for idx, refund in enumerate(refunds):
                    if idx in fallback_assigned_refund_indexes:
                        continue
                    if refund_has_known_member_id.get(idx, False):
                        continue
                    mededelingen = str(refund.get("mededelingen", ""))
                    refund_tegenrekening = normalize_account(refund.get("tegenrekening", ""))
                    if (
                        refund_match_ibans
                        and refund_tegenrekening
                        and refund_tegenrekening not in refund_match_ibans
                    ):
                        continue
                    if mededelingen and pattern.search(mededelingen):
                        refund_items.append({
                            "amount": round(abs(parse_amount(refund.get("amount", 0.0))), 2),
                            "mededelingen": mededelingen,
                            "tegenrekening": str(refund.get("tegenrekening", "")),
                            "index": idx,
                        })
                        fallback_assigned_refund_indexes.add(idx)
                        matched_refund_indexes.add(idx)
                        record["has_refund_transactions"] = True

        # Match previously processed orphaned refunds to known members by IBAN.
        if refund_match_ibans:
            matched_processed_notes = []
            existing_refund_keys = {
                build_refund_key(item.get("mededelingen", ""), item.get("amount", 0.0))
                for item in refund_items
            }
            for processed_entry in processed_with_meta:
                if processed_entry["used"]:
                    continue
                processed_iban = normalize_account(processed_entry.get("iban", ""))
                if not processed_iban or processed_iban not in refund_match_ibans:
                    continue

                processed_refund = processed_entry["refund"]
                processed_amount = round(abs(parse_amount(processed_refund.get("amount", 0.0))), 2)
                if processed_amount <= 0:
                    continue

                processed_key = build_refund_key(
                    processed_refund.get("mededelingen", ""),
                    processed_amount,
                )
                # Same refund already matched from bank transactions: consume processed entry, do not duplicate.
                if processed_key in existing_refund_keys:
                    processed_entry["used"] = True
                    continue

                refund_items.append(
                    {
                        "amount": processed_amount,
                        "mededelingen": str(processed_refund.get("mededelingen", "")),
                        "tegenrekening": processed_iban,
                    }
                )
                existing_refund_keys.add(processed_key)
                processed_entry["used"] = True
                record["has_refund_transactions"] = True

                reason = str(processed_refund.get("reason", "")).strip()
                if reason:
                    matched_processed_notes.append(f"Teruggestort € {processed_amount:.2f} - {reason}")
                else:
                    matched_processed_notes.append(f"Teruggestort € {processed_amount:.2f}")

            if matched_processed_notes:
                existing_note = str(record.get("penningmeester_opmerking", "")).strip()
                combined_note = " | ".join(matched_processed_notes)
                record["penningmeester_opmerking"] = (
                    f"{existing_note} | {combined_note}" if existing_note else combined_note
                )
        
        refunded_amount = round(sum(abs(parse_amount(item.get("amount", 0.0))) for item in refund_items), 2)
        record["refunded_amount"] = refunded_amount
        record["matched_refunds"] = refund_items
        
        # Check for manual refund overrides
        refund_override_entry = MANUAL_REFUND_OVERRIDES.get(member_id)
        if isinstance(refund_override_entry, dict) and refund_override_entry.get("amount"):
            manual_refund_amount = round(abs(parse_amount(refund_override_entry.get("amount", 0.0))), 2)
            refund_reason = str(refund_override_entry.get("reason", "")).strip()
            if manual_refund_amount > 0:
                refund_items.append({
                    "amount": manual_refund_amount,
                    "mededelingen": f"Handmatig teruggestort: {refund_reason}" if refund_reason else "Handmatig teruggestort",
                    "tegenrekening": "",
                })
                refunded_amount = round(refunded_amount + manual_refund_amount, 2)
                record["refunded_amount"] = refunded_amount
                record["matched_refunds"] = refund_items
                record["has_refund_transactions"] = True
                if refund_reason:
                    record["penningmeester_opmerking"] = (
                        f"Teruggestort € {manual_refund_amount:.2f} - {refund_reason}"
                    )

        # Ontvangen bedrag in de hoofdtabel is netto: ontvangen min teruggestort.
        net_received_amount = round(received_amount - abs(refunded_amount), 2)
        record["received_amount"] = net_received_amount

        tegenrekeningen = []
        seen_tegenrekeningen = set()
        for item in matched_items:
            tr = normalize_account(item.get("tegenrekening", ""))
            if tr and tr not in seen_tegenrekeningen:
                seen_tegenrekeningen.add(tr)
                tegenrekeningen.append(tr)
        record["tegenrekening"] = " | ".join(tegenrekeningen)

        match_indexes = {
            "all": all_matches,
            "id": matched_by_id,
            "manual": matched_by_manual,
            "fallback": matched_by_fallback,
            "iban": matched_by_iban,
        }

        if match_indexes["manual"]:
            record["opmerking"] = "Handmatig gematched via config"

        split_transactions_count = sum(
            1
            for item in matched_items
            if int(item.get("split_count", 1)) > 1
        )
        if split_transactions_count > 0:
            tx_label = "transactie" if split_transactions_count == 1 else "transacties"
            record["opmerking"] = (
                f"Bedrag verdeeld over meerdere lidnummers in {split_transactions_count} {tx_label}"
            )

        if match_indexes["fallback"]:
            fallback_count = len(match_indexes["fallback"])
            tx_label = "transactie" if fallback_count == 1 else "transacties"
            achternaam = record.get("achternaam", "")
            record["opmerking"] = (
                f"Geen 4-cijferig ID ({member_id_4digit}) gevonden; "
                f"backup match op achternaam '{achternaam}' ({fallback_count} {tx_label})"
            )

        if match_indexes["iban"]:
            iban_count = len(match_indexes["iban"])
            tx_label = "transactie" if iban_count == 1 else "transacties"
            record["opmerking"] = f"Backup match op tegenrekening (IBAN) ({iban_count} {tx_label})"

        matched_transaction_keys.update(all_indexes)

        if record.get("manual_paid_override"):
            reason = record.get("manual_paid_reason", "")
            if due_amount > 0 and net_received_amount < due_amount:
                record["received_amount"] = due_amount
                net_received_amount = due_amount

            record["opmerking"] = (
                f"Handmatig als betaald gemarkeerd: {reason}"
                if reason
                else "Handmatig als betaald gemarkeerd"
            )
            record["status_icon"] = "✅"
            record["status_label"] = "Handmatig gemarkeerd als betaald"
            record["status_class"] = "status-ok"
            exact_count += 1
            continue

        if abs(due_amount) < 0.005 and abs(net_received_amount) < 0.005:
            record["opmerking"] = "Geen contributie verschuldigd"
            record["status_icon"] = "🟣"
            record["status_label"] = "Niet van toepassing (0/0)"
            record["status_class"] = "status-neutral"
            neutral_count += 1
            continue

        # received_amount is above al netto gezet; gebruik die waarde direct.
        net_received_amount = round(record.get("received_amount", 0.0), 2)

        if abs(net_received_amount) < 0.005:
            if not match_indexes["fallback"]:
                record["opmerking"] = f"Geen 4-cijferig ID ({member_id_4digit}) gevonden in mededelingen"
                record["status_icon"] = "❌"
                record["status_label"] = "Nog niets ontvangen"
                record["status_class"] = "status-none"
                none_count += 1
                continue

        if abs(net_received_amount - due_amount) <= 0.009:
            record["status_icon"] = "✅"
            record["status_label"] = "Volledig ontvangen"
            record["status_class"] = "status-ok"
            exact_count += 1
            continue

        if net_received_amount < due_amount:
            record["status_icon"] = "🔵"
            record["status_label"] = "Gedeeltelijk ontvangen"
            record["status_class"] = "status-partial"
            partial_count += 1
            continue

        record["status_icon"] = "🟠"
        record["status_label"] = "Ontvangen (meer dan te innen)"
        record["status_class"] = "status-overpaid"
        exact_count += 1

    unmatched_transactions = []
    for i, tx in enumerate(transactions):
        if i not in matched_transaction_keys:
            unmatched_transactions.append(tx)

    normalized_processed_refund_keys = set()
    for saved_key, saved_payload in PROCESSED_ORPHANED_REFUNDS.items():
        key_text = str(saved_key or "")
        if "|" in key_text:
            key_note, key_amount = key_text.split("|", 1)
            normalized_processed_refund_keys.add(build_refund_key(key_note, key_amount))
        elif isinstance(saved_payload, dict):
            normalized_processed_refund_keys.add(
                build_refund_key(key_text, saved_payload.get("amount", 0.0))
            )

    # Build orphaned refunds (refunds that don't match any known member ID)
    orphaned_refunds = []
    for idx, refund in enumerate(refunds):
        if idx in matched_refund_indexes:
            continue
        mededelingen = str(refund.get("mededelingen", ""))
        id_tokens = extract_4digit_tokens(mededelingen)
        
        # Check if this refund is already matched to a known member
        is_matched = False
        for token in id_tokens:
            if token in refunds_by_member_id_4digit:
                for rec in records:
                    if rec.get("member_id_4digit") == token:
                        is_matched = True
                        break
            if is_matched:
                break
        
        if not is_matched:
            # Check if this refund has already been processed
            refund_key = build_refund_key(mededelingen, refund.get("amount", 0.0))
            if refund_key not in normalized_processed_refund_keys:
                orphaned_refunds.append({
                    "mededelingen": mededelingen,
                    "amount": round(abs(float(refund.get("amount", 0.0))), 2),
                    "tegenrekening": normalize_account(refund.get("tegenrekening", "")),
                })

    processed_orphaned_refunds = []
    processed_orphaned_total = 0.0
    for processed_entry in processed_with_meta:
        if processed_entry["used"]:
            continue
        processed_refund = processed_entry["refund"]
        processed_orphaned_refunds.append(processed_refund)
        processed_orphaned_total += round(abs(parse_amount(processed_refund.get("amount", 0.0))), 2)

    # Pair unmatched payments with processed orphaned refunds by IBAN and move them to main records.
    unmatched_with_meta = []
    for tx in unmatched_transactions:
        tx_iban = normalize_account(tx.get("tegenrekening", "")) or extract_iban_from_text(tx.get("mededelingen", ""))
        tx_name = extract_name_from_text(tx.get("mededelingen", ""))
        unmatched_with_meta.append(
            {
                "tx": tx,
                "iban": tx_iban,
                "name": normalize_name_for_match(tx_name),
                "used": False,
            }
        )

    for processed_entry in processed_with_meta:
        if processed_entry["used"]:
            continue
        refund_name = extract_name_from_text(processed_entry["refund"].get("mededelingen", ""))
        processed_entry["name"] = normalize_name_for_match(refund_name)

    for payment_entry in unmatched_with_meta:
        payment_iban = payment_entry.get("iban", "")
        if not payment_iban or payment_entry["used"]:
            continue

        matched_refund_entry = None
        for refund_entry in processed_with_meta:
            if refund_entry["used"]:
                continue
            if refund_entry.get("iban", "") != payment_iban:
                continue

            payment_name = payment_entry.get("name", "")
            refund_name = refund_entry.get("name", "")
            # If both names are available, require a name match in addition to IBAN.
            if payment_name and refund_name and payment_name != refund_name:
                continue

            if refund_entry.get("iban", "") == payment_iban:
                matched_refund_entry = refund_entry
                break

        if not matched_refund_entry:
            continue

        tx = payment_entry["tx"]
        refund = matched_refund_entry["refund"]
        payment_amount = round(parse_amount(tx.get("amount", 0.0)), 2)
        refund_amount = round(abs(parse_amount(refund.get("amount", 0.0))), 2)
        mededelingen = str(tx.get("mededelingen", ""))
        extracted_ids = sorted(extract_4digit_tokens(mededelingen))
        fallback_member_id = extracted_ids[0] if extracted_ids else "-"

        net_received_amount = round(payment_amount - refund_amount, 2)
        if abs(net_received_amount) < 0.005:
            status_icon = "🟣"
            status_label = "Niet van toepassing (0/0)"
            status_class = "status-neutral"
        elif net_received_amount > 0:
            status_icon = "🔵"
            status_label = "Gedeeltelijk ontvangen"
            status_class = "status-partial"
        else:
            status_icon = "✅"
            status_label = "Ontvangen (meer dan te innen)"
            status_class = "status-ok"

        records.append(
            {
                "member_id": fallback_member_id,
                "member_id_4digit": normalize_member_id_4digit(fallback_member_id),
                "achternaam": "Niet-lid (gekoppeld op IBAN)",
                "email": "",
                "rekeningnummer": payment_iban,
                "due_amount": round(abs(payment_amount), 2),
                "received_amount": payment_amount,
                "refunded_amount": refund_amount,
                "matched_transactions": [
                    {
                        "amount": payment_amount,
                        "mededelingen": mededelingen,
                        "tegenrekening": str(tx.get("tegenrekening", "")),
                        "split_count": 1,
                    }
                ],
                "matched_refunds": [
                    {
                        "amount": refund_amount,
                        "mededelingen": str(refund.get("mededelingen", "")),
                        "tegenrekening": payment_iban,
                        "split_count": 1,
                    }
                ],
                "tegenrekening": payment_iban,
                "opmerking": "Automatisch gekoppeld via IBAN tussen niet-gematchte betaling en verwerkte terugstorting",
                "penningmeester_opmerking": str(refund.get("penningmeester_opmerking", "")).strip(),
                "manual_paid_override": False,
                "has_refund_transactions": True,
                "is_overledene_opzegger": True,
                "status_icon": status_icon,
                "status_label": status_label,
                "status_class": status_class,
            }
        )

        payment_entry["used"] = True
        matched_refund_entry["used"] = True

    unmatched_transactions = [entry["tx"] for entry in unmatched_with_meta if not entry["used"]]
    processed_orphaned_refunds = [entry["refund"] for entry in processed_with_meta if not entry["used"]]
    processed_orphaned_total = round(sum(item.get("amount", 0.0) for item in processed_orphaned_refunds), 2)

    status_sort_order = {
        "status-none": 0,
        "status-partial": 1,
        "status-ok": 2,
        "status-overpaid": 3,
        "status-neutral": 4,
    }
    records.sort(
        key=lambda item: (
            1 if item.get("is_overledene_opzegger") else 0,
            status_sort_order.get(item.get("status_class", ""), 99),
            item.get("achternaam", "").lower(),
            item.get("member_id", ""),
        )
    )

    exact_count = sum(1 for item in records if item.get("status_class") in {"status-ok", "status-overpaid"})
    partial_count = sum(1 for item in records if item.get("status_class") == "status-partial")
    none_count = sum(1 for item in records if item.get("status_class") == "status-none")
    neutral_count = sum(1 for item in records if item.get("status_class") == "status-neutral")

    stats = {
        "total": len(records),
        "exact": exact_count,
        "partial": partial_count,
        "none": none_count,
        "neutral": neutral_count,
        "total_due": sum(item.get("due_amount", 0.0) for item in records),
        "total_received": sum(item.get("received_amount", 0.0) for item in records),
        "total_refunded": (
            sum(item.get("refunded_amount", 0.0) for item in records)
            + processed_orphaned_total
        ),
        "total_net_received": (
            sum(item.get("received_amount", 0.0) for item in records)
            - processed_orphaned_total
        ),
        "unmatched_count": len(unmatched_transactions),
        "unmatched_total": round(sum(tx.get("amount", 0.0) for tx in unmatched_transactions), 2),
        "ambiguous_count": len(pending_ambiguous_transactions),
        "ambiguous_total": round(sum(tx.get("amount", 0.0) for tx in pending_ambiguous_transactions), 2),
        "orphaned_refund_count": len(orphaned_refunds),
        "orphaned_refund_total": round(sum(ref.get("amount", 0.0) for ref in orphaned_refunds), 2),
        "processed_orphaned_refund_count": len(processed_orphaned_refunds),
        "processed_orphaned_refund_total": round(processed_orphaned_total, 2),
    }

    return _cache_set(
        cache_key,
        (
            records,
            stats,
            errors,
            unmatched_transactions,
            pending_ambiguous_transactions,
            orphaned_refunds,
            processed_orphaned_refunds,
        ),
    )


@app.route("/")
def index():
    (
        records,
        stats,
        errors,
        unmatched_transactions,
        pending_ambiguous_transactions,
        orphaned_refunds,
        processed_orphaned_refunds,
    ) = build_overview()
    current_date = datetime.now().strftime("%d-%m-%Y")
    current_user = os.getlogin()
    return render_template(
        "contributie.html",
        records=records,
        stats=stats,
        errors=errors,
        unmatched_transactions=unmatched_transactions,
        pending_ambiguous_transactions=pending_ambiguous_transactions,
        orphaned_refunds=orphaned_refunds,
        processed_orphaned_refunds=processed_orphaned_refunds,
        current_date=current_date,
        current_user=current_user,
        main_app_url=MAIN_APP_URL,
    )


@app.route("/quit", methods=["POST"])
def quit_app():
    """Sluit de applicatie af."""
    try:
        logging.info("Applicatie wordt afgesloten...")
        logging.info("=" * 70)

        response = jsonify({"success": True, "message": "Applicatie sluit af"})

        def shutdown_server():
            import time
            time.sleep(1)
            logging.info("Flask server wordt beeindigd...")
            os._exit(0)

        import threading
        shutdown_thread = threading.Thread(target=shutdown_server, daemon=True)
        shutdown_thread.start()

        return response, 200
    except Exception as exc:
        logging.error(f"Fout bij afsluiten applicatie: {str(exc)}")
        return jsonify({"success": False, "message": f"Fout: {str(exc)}"}), 500


@app.route("/save_manual_mapping", methods=["POST"])
def save_manual_mapping():
    """Slaat een handmatige mapping op van mededelingen fragment naar lid nummer."""
    try:
        data = request.get_json()
        member_id = str(data.get("member_id", "")).strip()
        mededelingen = str(data.get("mededelingen", "")).strip()

        if not member_id or not mededelingen:
            return jsonify({"success": False, "error": "Lid nummer en mededelingen zijn verplicht"}), 400

        # Laad config
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config = json.load(f)

        # Zorg ervoor dat de manual_transaction_mappings sectie bestaat
        if "contributie" not in config:
            config["contributie"] = {}
        if "manual_transaction_mappings" not in config["contributie"]:
            config["contributie"]["manual_transaction_mappings"] = {}

        # Voeg de mapping toe
        config["contributie"]["manual_transaction_mappings"][member_id] = mededelingen

        # Update runtime mapping direct, zodat herstart niet nodig is.
        MANUAL_TRANSACTION_MAPPINGS[member_id] = mededelingen

        # Schrijf config terug naar bestand
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=4, ensure_ascii=False)

        invalidate_runtime_cache()

        logging.info(f"Manual mapping opgeslagen: lid {member_id} -> '{mededelingen[:50]}...'")
        return jsonify({"success": True, "message": "Mapping opgeslagen"}), 200

    except Exception as exc:
        logging.error(f"Fout bij opslaan manual mapping: {str(exc)}")
        return jsonify({"success": False, "error": f"Fout: {str(exc)}"}), 500


@app.route("/save_paid_override", methods=["POST"])
def save_paid_override():
    """Sla handmatige betaalstatus op per lid met reden."""
    try:
        data = request.get_json() or {}
        member_id = str(data.get("member_id", "")).strip()
        marked_paid = bool(data.get("marked_paid", False))
        reason = str(data.get("reason", "")).strip()

        if not member_id:
            return jsonify({"success": False, "error": "Lid nummer is verplicht"}), 400

        if marked_paid and not reason:
            return jsonify({"success": False, "error": "Reden is verplicht bij handmatig betaald"}), 400

        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config_data = json.load(f)

        if "contributie" not in config_data:
            config_data["contributie"] = {}
        if "manual_paid_overrides" not in config_data["contributie"]:
            config_data["contributie"]["manual_paid_overrides"] = {}

        if marked_paid:
            config_data["contributie"]["manual_paid_overrides"][member_id] = {
                "marked_paid": True,
                "reason": reason,
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            MANUAL_PAID_OVERRIDES[member_id] = config_data["contributie"]["manual_paid_overrides"][member_id]
            message = "Lid handmatig als betaald gemarkeerd"
        else:
            config_data["contributie"]["manual_paid_overrides"].pop(member_id, None)
            MANUAL_PAID_OVERRIDES.pop(member_id, None)
            message = "Handmatige betaald-markering verwijderd"

        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config_data, f, indent=4, ensure_ascii=False)

        invalidate_runtime_cache()
        logging.info("Paid override bijgewerkt voor lid %s | marked_paid=%s", member_id, marked_paid)
        return jsonify({"success": True, "message": message}), 200

    except Exception as exc:
        logging.error("Fout bij opslaan paid override: %s", str(exc))
        return jsonify({"success": False, "error": f"Fout: {str(exc)}"}), 500


@app.route("/save_split_decision", methods=["POST"])
def save_split_decision():
    """Sla handmatig besluit op voor transacties met meerdere lidnummers."""
    try:
        data = request.get_json() or {}
        mededelingen = str(data.get("mededelingen", "")).strip()
        mode = str(data.get("mode", "")).strip().lower()
        member_id = str(data.get("member_id", "")).strip()

        if not mededelingen:
            return jsonify({"success": False, "error": "Mededelingen zijn verplicht"}), 400

        if mode not in {"split", "single_member"}:
            return jsonify({"success": False, "error": "Ongeldige keuze"}), 400

        if mode == "single_member" and not member_id:
            return jsonify({"success": False, "error": "Lid nummer is verplicht bij niet splitsen"}), 400

        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config_data = json.load(f)

        if "contributie" not in config_data:
            config_data["contributie"] = {}
        if "manual_split_decisions" not in config_data["contributie"]:
            config_data["contributie"]["manual_split_decisions"] = {}

        decision_payload = {"mode": mode}
        if mode == "single_member":
            decision_payload["member_id"] = member_id

        config_data["contributie"]["manual_split_decisions"][mededelingen] = decision_payload
        MANUAL_SPLIT_DECISIONS[mededelingen] = decision_payload

        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config_data, f, indent=4, ensure_ascii=False)

        invalidate_runtime_cache()
        logging.info("Split-besluit opgeslagen: mode=%s member_id=%s", mode, member_id)
        return jsonify({"success": True, "message": "Besluit opgeslagen"}), 200

    except Exception as exc:
        logging.error("Fout bij opslaan split-besluit: %s", str(exc))
        return jsonify({"success": False, "error": f"Fout: {str(exc)}"}), 500


@app.route("/save_refund_override", methods=["POST"])
def save_refund_override():
    """Sla handmatige terugstorting op per lid met bedrag en reden."""
    try:
        data = request.get_json() or {}
        member_id = str(data.get("member_id", "")).strip()
        amount = abs(parse_amount(data.get("amount", 0)))
        reason = str(data.get("reason", "")).strip()

        if not member_id:
            return jsonify({"success": False, "error": "Lid nummer is verplicht"}), 400

        if amount <= 0:
            return jsonify({"success": False, "error": "Bedrag moet groter dan 0 zijn"}), 400

        if not reason:
            return jsonify({"success": False, "error": "Reden is verplicht"}), 400

        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config_data = json.load(f)

        if "contributie" not in config_data:
            config_data["contributie"] = {}
        if "manual_refund_overrides" not in config_data["contributie"]:
            config_data["contributie"]["manual_refund_overrides"] = {}

        config_data["contributie"]["manual_refund_overrides"][member_id] = {
            "amount": round(amount, 2),
            "reason": reason,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        MANUAL_REFUND_OVERRIDES[member_id] = config_data["contributie"]["manual_refund_overrides"][member_id]

        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config_data, f, indent=4, ensure_ascii=False)

        invalidate_runtime_cache()
        logging.info("Refund override bijgewerkt voor lid %s | amount=%.2f", member_id, amount)
        return jsonify({"success": True, "message": "Terugstorting opgeslagen"}), 200

    except Exception as exc:
        logging.error("Fout bij opslaan refund override: %s", str(exc))
        return jsonify({"success": False, "error": f"Fout: {str(exc)}"}), 500


@app.route("/save_orphaned_refund", methods=["POST"])
def save_orphaned_refund():
    """Sla verwerking van een orphaned refund op."""
    try:
        data = request.get_json() or {}
        mededelingen = str(data.get("mededelingen", "")).strip()
        amount = abs(parse_amount(data.get("amount", 0)))
        reason = str(data.get("reason", "")).strip()

        if not mededelingen:
            return jsonify({"success": False, "error": "Mededelingen zijn verplicht"}), 400

        if amount <= 0:
            return jsonify({"success": False, "error": "Bedrag moet groter dan 0 zijn"}), 400

        if not reason:
            return jsonify({"success": False, "error": "Reden is verplicht"}), 400

        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config_data = json.load(f)

        if "contributie" not in config_data:
            config_data["contributie"] = {}
        if "processed_orphaned_refunds" not in config_data["contributie"]:
            config_data["contributie"]["processed_orphaned_refunds"] = {}

        refund_key = build_refund_key(mededelingen, amount)
        config_data["contributie"]["processed_orphaned_refunds"][refund_key] = {
            "amount": round(amount, 2),
            "reason": reason,
            "processed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        PROCESSED_ORPHANED_REFUNDS[refund_key] = config_data["contributie"]["processed_orphaned_refunds"][refund_key]

        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config_data, f, indent=4, ensure_ascii=False)

        invalidate_runtime_cache()
        logging.info("Orphaned refund verwerkt | bedrag=%.2f | reden=%s", amount, reason)
        return jsonify({"success": True, "message": "Terugstorting verwerkt"}), 200

    except Exception as exc:
        logging.error("Fout bij opslaan orphaned refund: %s", str(exc))
        return jsonify({"success": False, "error": f"Fout: {str(exc)}"}), 500


if __name__ == "__main__":
    app.run(debug=False, host="127.0.0.1", port=5004)
