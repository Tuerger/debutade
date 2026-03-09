"""
Transactie Toevoegen Debutade - Web Applicatie
================================================

Applicatie voor toevoegen van transacties naar Excel bankrekening bestandem.

Versie: 1.0
Datum: 2026-03-05
"""

from __future__ import annotations

from copy import copy
from datetime import datetime, date
import csv
import getpass
import json
import logging
import os
import re
import secrets
import sys
import shutil
import time
from io import StringIO
from typing import Any

from flask import Flask, jsonify, redirect, render_template, request, send_file, g
from openpyxl import load_workbook
from openpyxl.styles import Alignment


if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        import io

        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.getenv(
    "DEBUTADE_CONFIG",
    os.path.abspath(os.path.join(SCRIPT_DIR, "..", "config.json")),
)


def load_config(config_path: str) -> dict[str, Any]:
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Configuratiebestand niet gevonden: {config_path}")

    with open(config_path, "r", encoding="utf-8") as config_file:
        root_config = json.load(config_file)

    shared = root_config.get("shared", {})
    bank = root_config.get("bankrekening", {})
    transactie_toevoegen = root_config.get("transactietoevoegen", {})

    grootboek_directory = shared.get("grootboek_directory", "")

    bank_excel_file_name = (
        shared.get("bank_excel_file_name", "")
    )
    
    bank_sheets = bank.get("required_sheets") or [
        "Bankrekening",
        "Spaarrekening 1",
        "Spaarrekening 2",
    ]
    
    bank_sheet_name = bank.get("excel_sheet_name") or "Bankrekening"

    def build_path(file_name: str) -> str:
        if not file_name:
            return ""
        if os.path.isabs(file_name):
            return file_name
        if grootboek_directory:
            return os.path.join(grootboek_directory, file_name)
        return file_name

    return {
        "bank_excel_path": build_path(bank_excel_file_name),
        "bank_sheets": bank_sheets,
        "bank_sheet_name": bank_sheet_name,
        "csv_import_directory": (
            transactie_toevoegen.get("csv_import_directory")
            or shared.get("csv_import_directory")
            or os.path.join(os.path.expanduser("~"), "Downloads")
        ),
        "backup_directory": shared.get("backup_directory", os.path.join(SCRIPT_DIR, "backups")),
        "log_directory": shared.get("log_directory", os.path.join(SCRIPT_DIR, "logs")),
        "log_level": shared.get("log_level", "INFO"),
        "main_app_url": os.getenv("MAIN_APP_URL", "").strip(),
    }


try:
    config = load_config(CONFIG_PATH)
except (FileNotFoundError, KeyError, json.JSONDecodeError) as exc:
    print(f"WAARSCHUWING: {exc}")
    config = {
        "bank_excel_path": "",
        "bank_sheets": ["Bankrekening", "Spaarrekening 1", "Spaarrekening 2"],
        "bank_sheet_name": "Bankrekening",
        "csv_import_directory": os.path.join(os.path.expanduser("~"), "Downloads"),
        "backup_directory": os.path.join(SCRIPT_DIR, "backups"),
        "log_directory": os.path.join(SCRIPT_DIR, "logs"),
        "log_level": "INFO",
        "main_app_url": os.getenv("MAIN_APP_URL", "").strip(),
    }


LOG_DIRECTORY = config["log_directory"]
LOG_LEVEL = config["log_level"]
MAIN_APP_URL = config["main_app_url"]

app = Flask(
    __name__,
    template_folder=os.path.join(SCRIPT_DIR, "templates"),
    static_folder=os.path.join(SCRIPT_DIR, "static"),
)
app.config["TEMPLATES_AUTO_RELOAD"] = True

CACHE_TTL_SECONDS = int(os.getenv("DEBUTADE_CACHE_TTL_SECONDS", "20"))
BENCHMARK_ENABLED = os.getenv("DEBUTADE_BENCHMARK", "1") == "1"
RUNTIME_CACHE: dict[tuple[Any, ...], dict[str, Any]] = {}
CSV_LOAD_TOKENS: dict[str, dict[str, Any]] = {}
CSV_LOAD_TOKEN_TTL_SECONDS = int(os.getenv("DEBUTADE_CSV_LOAD_TOKEN_TTL_SECONDS", "3600"))


def _file_signature(file_path: str) -> tuple[int, int] | None:
    if not file_path or not os.path.exists(file_path):
        return None
    stat = os.stat(file_path)
    return (stat.st_mtime_ns, stat.st_size)


def _cache_get(key: tuple[Any, ...]) -> Any:
    entry = RUNTIME_CACHE.get(key)
    if not entry:
        return None

    if time.time() - entry["ts"] > CACHE_TTL_SECONDS:
        RUNTIME_CACHE.pop(key, None)
        return None

    return entry["value"]


def _cache_set(key: tuple[Any, ...], value: Any) -> Any:
    RUNTIME_CACHE[key] = {"ts": time.time(), "value": value}
    return value


def invalidate_runtime_cache() -> None:
    RUNTIME_CACHE.clear()


@app.before_request
def _benchmark_start() -> None:
    if BENCHMARK_ENABLED:
        g._start_time = time.perf_counter()


@app.after_request
def _benchmark_end(response):
    if BENCHMARK_ENABLED and hasattr(g, "_start_time"):
        elapsed_ms = (time.perf_counter() - g._start_time) * 1000
        logging.info("PERF %s %s %s %.1fms", request.method, request.path, response.status_code, elapsed_ms)
    return response


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def find_header_column(col_map: dict[str, int], aliases: tuple[str, ...], default: int) -> int:
    """Return first matching column index from header aliases, else fallback to default."""
    for alias in aliases:
        idx = col_map.get(alias)
        if idx is not None:
            return idx
    return default


def parse_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if value is None:
        return None

    raw = normalize_text(value)
    if not raw:
        return None

    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%y", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def parse_amount(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)

    raw = normalize_text(value)
    if not raw:
        return None

    raw = raw.replace("€", "").replace(" ", "")
    if raw.count(",") == 1 and raw.count(".") > 1:
        raw = raw.replace(".", "")
    raw = raw.replace(".", "").replace(",", ".")

    try:
        return float(raw)
    except ValueError:
        return None


def extract_valutadatum_from_mededelingen(mededelingen: str) -> str | None:
    """Extract valutadatum from bank mededelingen field (format: 'Valutadatum: DD-MM-YYYY')"""
    if not mededelingen:
        return None
    
    # Look for pattern "Valutadatum: DD-MM-YYYY"
    match = re.search(r"Valutadatum:\s*(\d{2}-\d{2}-\d{4})", mededelingen, re.IGNORECASE)
    if match:
        return match.group(1)
    
    return None


def read_transactions_from_sheet(file_path: str, sheet_name: str) -> tuple[list[dict[str, Any]], str | None]:
    """Read transactions from sheet. Returns (records, error_message)"""
    records: list[dict[str, Any]] = []
    workbook = None

    signature = _file_signature(file_path)
    cache_key = ("sheet_records", file_path, sheet_name, signature)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    if not file_path:
        return records, "Bestandspad niet ingesteld"
    
    if not os.path.exists(file_path):
        return records, f"Bestand niet gevonden: {os.path.basename(file_path)}"

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            return records, f"Sheet '{sheet_name}' niet gevonden in bestand"

        sheet = workbook[sheet_name]

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_map: dict[str, int] = {}
        if header_row:
            for index, cell_value in enumerate(header_row):
                key = normalize_text(cell_value).lower()
                if key and key not in header_map:
                    header_map[key] = index

        def get_value(row: tuple[Any, ...], header_names: tuple[str, ...], fallback_index: int | None = None) -> Any:
            for header_name in header_names:
                idx = header_map.get(header_name.lower())
                if idx is not None and idx < len(row):
                    return row[idx]
            if fallback_index is not None and fallback_index < len(row):
                return row[fallback_index]
            return None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            txn_date = parse_date(get_value(row, ("datum",), 0))
            af_bij = normalize_text(get_value(row, ("af bij", "af/bij"), 5))
            amount = parse_amount(get_value(row, ("bedrag (eur)", "bedrag", "amount"), 6))
            rekening = normalize_text(get_value(row, ("rekening",), 2))
            tegenrekening = normalize_text(get_value(row, ("tegenrekening",), 3))
            code = normalize_text(get_value(row, ("code",), 4))
            valutadatum = normalize_text(get_value(row, ("valutadatum",), None))
            mededelingen = normalize_text(get_value(row, ("mededelingen",), 8))
            tag = normalize_text(get_value(row, ("tag",), 11))

            if amount is None or af_bij not in {"Af", "Bij"}:
                continue

            records.append(
                {
                    "date": txn_date.strftime("%d-%m-%Y") if txn_date else "",
                    "datum_object": txn_date,
                    "naam_omschrijving": normalize_text(get_value(row, ("naam / omschrijving", "omschrijving", "naam"), 1)),
                    "af_bij": af_bij,
                    "bedrag": round(float(amount), 2),
                    "rekening": rekening,
                    "tegenrekening": tegenrekening,
                    "code": code,
                    "valutadatum": valutadatum,
                    "mededelingen": mededelingen,
                    "tag": tag,
                }
            )

        records.sort(key=lambda item: (item["datum_object"] or datetime(1900, 1, 1)), reverse=True)
        return _cache_set(cache_key, (records, None))
    except PermissionError as exc:
        error_msg = f"Kan bestand niet openen: {os.path.basename(file_path)}. Het bestand is mogelijk nog geopend in Excel of wordt gesynchroniseerd."
        logging.error("Permission denied bij lezen sheet %s uit %s", sheet_name, file_path)
        return records, error_msg
    except Exception as exc:
        error_msg = f"Fout bij lezen sheet '{sheet_name}': {str(exc)}"
        logging.error("Fout bij lezen sheet %s uit %s: %s", sheet_name, file_path, exc)
        return records, error_msg
    finally:
        if workbook:
            workbook.close()


def get_sheet_stats(file_path: str, sheet_name: str) -> tuple[dict[str, Any], str | None]:
    """Get statistics for a sheet. Returns (stats, error_message)"""
    records, error = read_transactions_from_sheet(file_path, sheet_name)
    
    if error:
        return {"count": 0, "last_date": None}, error
    
    last_date = None
    record_count = len(records)
    
    if records and records[0].get("datum_object"):
        last_date = records[0]["datum_object"].strftime("%d-%m-%Y")
    
    return {"count": record_count, "last_date": last_date}, None


def get_all_existing_transactions(file_path: str, sheet_names: list[str]) -> tuple[dict[str, set], str | None]:
    """Get all existing transactions for duplicate detection. Returns (existing_txns, error_message)"""
    signature = _file_signature(file_path)
    cache_key = ("existing_txns", file_path, tuple(sheet_names), signature)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    existing_txns = {}
    
    for sheet_name in sheet_names:
        records, error = read_transactions_from_sheet(file_path, sheet_name)
        if error and not records:
            # Return error on first sheet that fails
            return {}, error
        # Create a tuple key: (date, amount, description)
        txn_set = {
            (r["date"], str(r["bedrag"]), r["naam_omschrijving"]) 
            for r in records
        }
        existing_txns[sheet_name] = txn_set
    
    return _cache_set(cache_key, (existing_txns, None))


def get_valid_recheningen() -> dict[str, str]:
    """Get valid bank account numbers for each sheet"""
    return {
        "Bankrekening": "NL40INGB0002691632",
        "Spaarrekening 1": "S 858-17363",
        "Spaarrekening 2": "D 130-41072",
    }


def validate_transaction(txn: dict[str, Any], sheet_name: str) -> tuple[bool, list[str]]:
    """Validate a transaction and return (is_valid, error_messages)"""
    errors = []
    
    # Datum validation
    if not txn.get("date"):
        errors.append(f"Datum ontbreekt: '{txn.get('date')}'")
    
    # Bedrag validation
    if txn.get("bedrag") is None or not isinstance(txn.get("bedrag"), (int, float)):
        errors.append(f"Bedrag moet numeriek zijn: '{txn.get('bedrag')}'")
    
    # Af Bij validation
    if txn.get("af_bij") not in {"Af", "Bij"}:
        errors.append(f"Af Bij mag alleen 'Af' of 'Bij' zijn: '{txn.get('af_bij')}'")
    
    # Rekening validation
    valid_rekeningen = get_valid_recheningen()
    if sheet_name in valid_rekeningen:
        expected_rekening = valid_rekeningen[sheet_name]
        if txn.get("rekening") != expected_rekening:
            errors.append(
                f"Rekeningnummer incorrect voor {sheet_name}. "
                f"Verwacht: '{expected_rekening}', Gekregen: '{txn.get('rekening')}'"
            )
    
    return len(errors) == 0, errors


def find_column_value(csv_row: dict[str, Any], possible_names: tuple[str, ...]) -> str:
    """Find a column value by trying multiple possible names (case-insensitive)."""
    name_map = {normalize_text(col).lower(): col for col in csv_row.keys()}
    for candidate in possible_names:
        key = normalize_text(candidate).lower()
        col_name = name_map.get(key)
        if col_name is not None:
            return csv_row.get(col_name, "")
    return ""


def parse_csv_rows(sheet_name: str, csv_data: list[dict[str, Any]]) -> tuple[dict[str, Any], int]:
    """Build parse response payload from CSV rows for given sheet."""
    bank_file = config.get("bank_excel_path", "")
    existing_txns, error = get_all_existing_transactions(bank_file, config.get("bank_sheets", []))

    if error:
        return {
            "error": error,
            "transactions": [],
            "validation_errors": [],
            "duplicates": [],
        }, 500

    existing_in_sheet = existing_txns.get(sheet_name, set())
    parsed_transactions = []
    validation_errors = []
    duplicates = []

    for idx, row in enumerate(csv_data, start=1):
        datum_val = find_column_value(row, ("Datum", "datum", "Date", "date"))
        bedrag_val = find_column_value(row, ("Bedrag", "bedrag", "Amount", "amount", "Bedrag (EUR)", "bedrag (eur)"))
        af_bij_val = find_column_value(row, ("Af Bij", "af_bij", "Af/Bij", "af/bij"))
        naam_val = find_column_value(row, ("Naam / Omschrijving", "Omschrijving", "Naam", "naam", "omschrijving", "Description", "description"))
        rekening_val = find_column_value(row, ("Rekening", "rekening", "IBAN", "iban", "Account"))
        tegenrekening_val = find_column_value(row, ("Tegenrekening", "tegenrekening", "Counter Account", "counter account"))
        code_val = find_column_value(row, ("Code", "code", "Mutatiesoort", "mutatiesoort", "Type"))
        valutadatum_val = find_column_value(row, ("Valutadatum", "valutadatum", "Value Date", "valuedate"))
        mededelingen_val = find_column_value(row, ("Mededelingen", "mededelingen", "Notes", "notes", "Opmerking"))
        tag_val = find_column_value(row, ("Tag", "tag", "Tags", "tags"))

        txn_date = parse_date(datum_val)
        bedrag = parse_amount(bedrag_val)
        af_bij = normalize_text(af_bij_val)

        valutadatum_extracted = valutadatum_val
        if not valutadatum_extracted and mededelingen_val:
            valutadatum_extracted = extract_valutadatum_from_mededelingen(mededelingen_val) or ""

        txn = {
            "date": txn_date.strftime("%d-%m-%Y") if txn_date else "",
            "datum_object": txn_date,
            "naam_omschrijving": normalize_text(naam_val),
            "af_bij": af_bij,
            "bedrag": bedrag,
            "rekening": normalize_text(rekening_val),
            "tegenrekening": normalize_text(tegenrekening_val),
            "code": normalize_text(code_val),
            "valutadatum": normalize_text(valutadatum_extracted),
            "mededelingen": normalize_text(mededelingen_val),
            "tag": normalize_text(tag_val),
            "row_index": idx,
        }

        is_valid, errors = validate_transaction(txn, sheet_name)
        if not is_valid:
            for err in errors:
                validation_errors.append(f"Rij {idx}: {err}")
        else:
            txn_key = (txn["date"], str(txn["bedrag"]), txn["naam_omschrijving"])
            if txn_key in existing_in_sheet:
                duplicates.append(
                    f"Rij {idx}: Transaction al aanwezig (Datum: {txn['date']}, Bedrag: {txn['bedrag']}, Beschrijving: {txn['naam_omschrijving']})"
                )
            else:
                parsed_transactions.append(txn)

    parsed_transactions.sort(key=lambda item: (item["datum_object"] or datetime(1900, 1, 1)), reverse=True)

    if validation_errors or duplicates:
        return {
            "valid": False,
            "transactions": parsed_transactions,
            "validation_errors": validation_errors,
            "duplicates": duplicates,
        }, 200

    if not parsed_transactions:
        return {
            "valid": False,
            "transactions": [],
            "validation_errors": ["Geen geldige transacties gevonden in bestand"],
            "duplicates": [],
        }, 200

    return {
        "valid": True,
        "transactions": parsed_transactions,
        "validation_errors": [],
        "duplicates": [],
    }, 200


def resolve_csv_directory() -> str:
    configured = str(config.get("csv_import_directory", "") or "").strip()
    if configured:
        return os.path.abspath(configured)
    return os.path.abspath(os.path.join(os.path.expanduser("~"), "Downloads"))


def get_dropdown_csv_entries() -> list[dict[str, Any]]:
    """Return the same CSV list (sorting + cap) as shown in the dropdown."""
    csv_dir = resolve_csv_directory()
    if not os.path.isdir(csv_dir):
        return []

    files = []
    for name in os.listdir(csv_dir):
        if not name.lower().endswith(".csv"):
            continue
        full_path = os.path.join(csv_dir, name)
        if not os.path.isfile(full_path):
            continue
        stat = os.stat(full_path)
        files.append(
            {
                "name": name,
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "mtime": stat.st_mtime,
            }
        )

    files.sort(key=lambda item: item["mtime"], reverse=True)
    return files[:200]


def store_csv_load_token(source_file_name: str, backup_path: str) -> str:
    token = secrets.token_urlsafe(24)
    CSV_LOAD_TOKENS[token] = {
        "ts": time.time(),
        "source_file_name": source_file_name,
        "backup_path": backup_path,
    }
    return token


def consume_csv_load_token(token: str) -> dict[str, Any] | None:
    token_data = CSV_LOAD_TOKENS.pop(token, None)
    if not token_data:
        return None

    if time.time() - token_data.get("ts", 0) > CSV_LOAD_TOKEN_TTL_SECONDS:
        return None
    return token_data


def backup_loaded_csv(csv_path: str) -> tuple[bool, str]:
    """Copy selected CSV to backup directory with a timestamp suffix."""
    backup_dir = str(config.get("backup_directory", "") or "").strip()
    if not backup_dir:
        return False, "Backup directory niet geconfigureerd"

    try:
        os.makedirs(backup_dir, exist_ok=True)
    except Exception as exc:
        return False, f"Kan backup directory niet aanmaken: {exc}"

    base_name = os.path.basename(csv_path)
    name, ext = os.path.splitext(base_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target_name = f"{name}_loaded_{timestamp}{ext}"
    target_path = os.path.join(backup_dir, target_name)

    try:
        shutil.copy2(csv_path, target_path)
        return True, target_path
    except Exception as exc:
        return False, f"Kon CSV backup niet maken: {exc}"


def maybe_remove_loaded_csv(source_file_name: str, csv_backup_path: str, expected_txn_count: int, added_count: int) -> tuple[bool, str]:
    """Remove source CSV only if all transactions were added and backup copy is valid."""
    if added_count != expected_txn_count:
        return False, "CSV niet verwijderd: niet alle transacties zijn toegevoegd"

    safe_name = os.path.basename(source_file_name or "")
    if not safe_name or safe_name != (source_file_name or ""):
        return False, "CSV niet verwijderd: ongeldige bestandsnaam"

    csv_dir = resolve_csv_directory()
    source_path = os.path.abspath(os.path.join(csv_dir, safe_name))
    if os.path.dirname(source_path) != csv_dir:
        return False, "CSV niet verwijderd: bestand buiten toegestane map"

    backup_dir = str(config.get("backup_directory", "") or "").strip()
    if not backup_dir:
        return False, "CSV niet verwijderd: backup directory niet geconfigureerd"

    backup_dir_abs = os.path.abspath(backup_dir)
    backup_path_abs = os.path.abspath(csv_backup_path or "")
    if not backup_path_abs.startswith(backup_dir_abs + os.sep):
        return False, "CSV niet verwijderd: backup pad is ongeldig"
    if not os.path.exists(backup_path_abs):
        return False, "CSV niet verwijderd: backup bestand ontbreekt"

    source_ext = os.path.splitext(safe_name)[1].lower()
    backup_name = os.path.basename(backup_path_abs)
    if source_ext != ".csv" or not backup_name.lower().endswith(".csv"):
        return False, "CSV niet verwijderd: bestandstype mismatch"

    if not os.path.exists(source_path):
        return False, "CSV al verwijderd of niet meer aanwezig"

    try:
        os.remove(source_path)
        return True, source_path
    except Exception as exc:
        return False, f"CSV niet verwijderd: {exc}"


@app.before_request
def log_request() -> None:
    logging.info("REQUEST %s %s %s", request.remote_addr, request.method, request.path)


@app.route("/")
def index():
    bank_file = config.get("bank_excel_path", "")
    
    stats = {}
    error = None
    for sheet_name in config.get("bank_sheets", []):
        stat, err = get_sheet_stats(bank_file, sheet_name)
        stats[sheet_name] = stat
        if err and not error:
            error = err
    
    return render_template(
        "index.html",
        current_date=datetime.now().strftime("%d-%m-%Y"),
        current_user=getpass.getuser(),
        main_app_url=MAIN_APP_URL,
        bank_sheets=config.get("bank_sheets", []),
        csv_import_directory=resolve_csv_directory(),
        stats=stats,
        error=error,
    )


@app.route("/api/csv-files")
def csv_files():
    csv_dir = resolve_csv_directory()
    if not os.path.isdir(csv_dir):
        return jsonify({
            "directory": csv_dir,
            "files": [],
            "error": f"CSV map niet gevonden: {csv_dir}",
        }), 200

    try:
        files = get_dropdown_csv_entries()
        for item in files:
            item.pop("mtime", None)
        return jsonify({"directory": csv_dir, "files": files})
    except Exception as exc:
        logging.error("Fout bij ophalen CSV bestanden uit %s: %s", csv_dir, exc)
        return jsonify({
            "directory": csv_dir,
            "files": [],
            "error": f"Fout bij lezen CSV map: {exc}",
        }), 500


@app.route("/api/sheet-data/<sheet_name>")
def sheet_data(sheet_name: str):
    bank_file = config.get("bank_excel_path", "")
    bank_sheets = config.get("bank_sheets", [])
    
    if sheet_name not in bank_sheets:
        return jsonify({"error": "Sheet niet geldig"}), 400
    
    if not bank_file or not os.path.exists(bank_file):
        return jsonify({"error": f"Kan bestand niet openen: {os.path.basename(bank_file) if bank_file else 'onbekend'}"}), 400
    
    records, error = read_transactions_from_sheet(bank_file, sheet_name)
    
    if error:
        return jsonify({"error": error}), 500
    
    stats, _ = get_sheet_stats(bank_file, sheet_name)
    
    return jsonify({
        "transactions": records,
        "stats": stats,
    })


@app.route("/api/parse-csv", methods=["POST"])
def parse_csv():
    """Parse uploaded CSV and return preview with validation"""
    if "file" not in request.files:
        return jsonify({"error": "Geen bestand gerupload"}), 400
    
    file = request.files["file"]
    sheet_name = request.form.get("sheet_name")
    
    if not sheet_name:
        return jsonify({"error": "Sheet naam niet opgegeven"}), 400
    
    if sheet_name not in config.get("bank_sheets", []):
        return jsonify({"error": "Sheet naam niet geldig"}), 400
    
    if not file or not file.filename.endswith(".csv"):
        return jsonify({"error": "Bestand moet CSV zijn"}), 400
    
    try:
        stream = StringIO(file.stream.read().decode("utf-8-sig", errors="replace"), newline=None)
        csv_data = list(csv.DictReader(stream, delimiter=';'))
    except Exception as exc:
        return jsonify({"error": f"CSV parsing fout: {str(exc)}"}), 400
    payload, status = parse_csv_rows(sheet_name, csv_data)
    return jsonify(payload), status


@app.route("/api/parse-csv-path", methods=["POST"])
def parse_csv_path():
    data = request.get_json() or {}
    sheet_name = normalize_text(data.get("sheet_name"))
    file_name = normalize_text(data.get("file_name"))

    if not sheet_name:
        return jsonify({"error": "Sheet naam niet opgegeven"}), 400
    if sheet_name not in config.get("bank_sheets", []):
        return jsonify({"error": "Sheet naam niet geldig"}), 400
    if not file_name:
        return jsonify({"error": "Bestandsnaam ontbreekt"}), 400

    csv_dir = resolve_csv_directory()
    safe_name = os.path.basename(file_name)
    if safe_name != file_name or not safe_name.lower().endswith(".csv"):
        return jsonify({"error": "Ongeldige bestandsnaam"}), 400

    dropdown_entries = get_dropdown_csv_entries()
    dropdown_names = {entry["name"] for entry in dropdown_entries}
    if safe_name not in dropdown_names:
        return jsonify({"error": "Bestand staat niet in actuele CSV dropdownlijst"}), 400

    full_path = os.path.abspath(os.path.join(csv_dir, safe_name))
    if os.path.dirname(full_path) != csv_dir:
        return jsonify({"error": "Bestand buiten toegestane map"}), 400
    if not os.path.exists(full_path):
        return jsonify({"error": f"Bestand niet gevonden: {safe_name}"}), 404

    try:
        with open(full_path, "r", encoding="utf-8-sig", errors="replace", newline=None) as handle:
            csv_data = list(csv.DictReader(handle, delimiter=';'))
    except Exception as exc:
        return jsonify({"error": f"CSV parsing fout: {exc}"}), 400

    copied, backup_info = backup_loaded_csv(full_path)
    if copied:
        logging.info("CSV backup gemaakt: %s", backup_info)
    else:
        logging.warning("CSV backup overgeslagen/mislukt voor %s: %s", full_path, backup_info)

    payload, status = parse_csv_rows(sheet_name, csv_data)
    csv_removed = False
    csv_remove_message = ""

    # Als alle records al aanwezig zijn (alleen duplicaten, geen validatiefouten),
    # mag de bron-CSV direct weg nadat backup is bevestigd.
    has_only_duplicates = (
        bool(payload.get("duplicates"))
        and not payload.get("transactions")
        and not payload.get("validation_errors")
    )
    load_token = ""
    if copied:
        load_token = store_csv_load_token(safe_name, backup_info)

    if copied and has_only_duplicates:
        csv_removed, csv_remove_message = maybe_remove_loaded_csv(
            source_file_name=safe_name,
            csv_backup_path=backup_info,
            expected_txn_count=0,
            added_count=0,
        )
        if csv_removed:
            logging.info("Bron CSV verwijderd (alleen duplicaten): %s", csv_remove_message)
        else:
            logging.warning("Bron CSV niet verwijderd (alleen duplicaten): %s", csv_remove_message)

    if copied:
        payload["csv_backup_path"] = backup_info
        payload["source_file_token"] = load_token
    else:
        payload["csv_backup_warning"] = backup_info
    payload["csv_removed"] = csv_removed
    payload["csv_remove_message"] = csv_remove_message
    return jsonify(payload), status


def sort_sheet_by_date(excel_path, sheet_name, datum_col):
    """Sort all transactions in sheet by date (newest first)"""
    workbook = load_workbook(excel_path)
    sheet = workbook[sheet_name]

    # Read row values + styles (skip header at row 1)
    max_col = sheet.max_column
    all_rows = []
    for row in sheet.iter_rows(min_row=2, max_col=max_col):
        row_snapshot = []
        for cell in row:
            row_snapshot.append({
                "value": cell.value,
                "style": copy(cell._style),
            })
        all_rows.append(row_snapshot)

    # Parse dates and sort
    def get_sort_date(row_snapshot):
        try:
            date_value = row_snapshot[datum_col - 1]["value"] if len(row_snapshot) >= datum_col else ""
            if date_value:
                parsed = parse_date(date_value)
                return parsed if parsed else datetime(1900, 1, 1)
            return datetime(1900, 1, 1)
        except Exception:
            return datetime(1900, 1, 1)

    # Sort newest first
    all_rows.sort(key=get_sort_date, reverse=True)

    # Write sorted rows back with original styles preserved
    for new_row_idx, row_snapshot in enumerate(all_rows, start=2):
        for col_idx, cell_data in enumerate(row_snapshot, start=1):
            cell = sheet.cell(row=new_row_idx, column=col_idx)
            cell.value = cell_data["value"]
            cell._style = copy(cell_data["style"])
    
    workbook.save(excel_path)
    workbook.close()
    logging.info(f"Sheet {sheet_name} gesorteerd op datum (nieuwste eerst)")


@app.route("/api/add-transactions", methods=["POST"])
def add_transactions():
    """Add validated transactions to Excel"""
    data = request.get_json()
    sheet_name = data.get("sheet_name")
    transactions = data.get("transactions", [])
    source_file_name = normalize_text(data.get("source_file_name"))
    csv_backup_path = normalize_text(data.get("csv_backup_path"))
    source_file_token = normalize_text(data.get("source_file_token"))
    
    if not sheet_name or sheet_name not in config.get("bank_sheets", []):
        return jsonify({"error": "Sheet naam niet geldig"}), 400
    
    if not transactions:
        return jsonify({"error": "Geen transacties opgegeven"}), 400
    
    bank_file = config.get("bank_excel_path", "")
    backup_dir = config.get("backup_directory", "")
    
    if not bank_file or not os.path.exists(bank_file):
        return jsonify({"error": "Bank bestand niet gevonden"}), 400
    
    try:
        # Create backup
        if backup_dir and not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        backup_name = f"Debutade boekjaar bank 2026_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        backup_path = os.path.join(backup_dir, backup_name)
        shutil.copy2(bank_file, backup_path)
        logging.info(f"Backup gemaakt: {backup_path}")
        
        # Open workbook and determine column mapping
        workbook = load_workbook(bank_file)
        
        if sheet_name not in workbook.sheetnames:
            return jsonify({"error": f"Sheet {sheet_name} niet gevonden in werkboek"}), 400
        
        sheet = workbook[sheet_name]
        
        # Get header mapping from existing headers
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        col_map = {}
        if header_row:
            for idx, cell_value in enumerate(header_row):
                key = normalize_text(cell_value).lower()
                col_map[key] = idx + 1  # openpyxl uses 1-based indexing
        
        # Find column positions by aliases; defaults match the standard bank sheet layout.
        datum_col = find_header_column(col_map, ("datum",), 1)
        naam_col = find_header_column(col_map, ("naam / omschrijving", "omschrijving", "naam"), 2)
        rekening_col = find_header_column(col_map, ("rekening",), 3)
        tegenrekening_col = find_header_column(col_map, ("tegenrekening",), 4)
        code_col = find_header_column(col_map, ("code",), 5)
        af_bij_col = find_header_column(col_map, ("af bij", "af/bij"), 6)
        bedrag_col = find_header_column(col_map, ("bedrag (eur)", "bedrag", "amount"), 7)
        valuta_col = find_header_column(col_map, ("valutadatum",), 8)
        mededelingen_col = find_header_column(col_map, ("mededelingen",), 9)
        tag_col = find_header_column(col_map, ("tag",), 12)
        
        # Find next row
        next_row = sheet.max_row + 1
        
        added_count = 0
        for txn in transactions:
            # Write a true Excel date value (datetime), not a formatted text string.
            date_value = parse_date(txn.get("date", ""))

            date_cell = sheet.cell(row=next_row, column=datum_col)
            date_cell.value = date_value
            if date_value:
                date_cell.number_format = "DD-MM-YYYY"
            date_cell.alignment = Alignment(horizontal="right")

            sheet.cell(row=next_row, column=naam_col).value = txn.get("naam_omschrijving")
            sheet.cell(row=next_row, column=rekening_col).value = txn.get("rekening")
            sheet.cell(row=next_row, column=tegenrekening_col).value = txn.get("tegenrekening", "")
            sheet.cell(row=next_row, column=code_col).value = txn.get("code", "")
            sheet.cell(row=next_row, column=valuta_col).value = txn.get("valutadatum")
            sheet.cell(row=next_row, column=mededelingen_col).value = txn.get("mededelingen")
            sheet.cell(row=next_row, column=af_bij_col).value = txn.get("af_bij")
            sheet.cell(row=next_row, column=bedrag_col).value = txn.get("bedrag")
            sheet.cell(row=next_row, column=tag_col).value = txn.get("tag", "")
            # Saldo na mutatie (if it exists) wordt niet ingevuld per requirements
            
            next_row += 1
            added_count += 1
        
        workbook.save(bank_file)
        workbook.close()
        invalidate_runtime_cache()
        
        logging.info(f"{added_count} transacties toegevoegd aan {sheet_name}")
        
        # Sort sheet by date (newest first)
        try:
            sort_sheet_by_date(bank_file, sheet_name, datum_col)
        except Exception as sort_exc:
            logging.warning(f"Kon sheet niet sorteren: {str(sort_exc)}")
        
        # Get updated stats
        stats, _ = get_sheet_stats(bank_file, sheet_name)
        
        csv_removed = False
        csv_remove_message = ""
        if source_file_name and csv_backup_path and source_file_token:
            token_data = consume_csv_load_token(source_file_token)
            if not token_data:
                csv_removed = False
                csv_remove_message = "CSV niet verwijderd: ongeldig of verlopen laadtoken"
            elif token_data.get("source_file_name") != source_file_name or token_data.get("backup_path") != csv_backup_path:
                csv_removed = False
                csv_remove_message = "CSV niet verwijderd: laadtoken komt niet overeen met bestand/backup"
            else:
                csv_removed, csv_remove_message = maybe_remove_loaded_csv(
                    source_file_name=source_file_name,
                    csv_backup_path=csv_backup_path,
                    expected_txn_count=len(transactions),
                    added_count=added_count,
                )
            if csv_removed:
                logging.info("Bron CSV verwijderd na succesvolle import: %s", csv_remove_message)
            else:
                logging.warning("Bron CSV niet verwijderd: %s", csv_remove_message)
        elif source_file_name or csv_backup_path:
            csv_removed = False
            csv_remove_message = "CSV niet verwijderd: ontbrekend laadtoken of backup informatie"

        return jsonify({
            "success": True,
            "added_count": added_count,
            "stats": stats,
            "message": f"{added_count} transacties succesvol toegevoegd",
            "csv_removed": csv_removed,
            "csv_remove_message": csv_remove_message,
        })
    
    except Exception as exc:
        logging.error(f"Fout bij toevoegen transacties: {str(exc)}")
        return jsonify({"error": f"Fout bij toevoegen: {str(exc)}"}), 500


@app.route("/quit", methods=["POST"])
def quit_application():
    try:
        user = getpass.getuser()
        logging.info("APPLICATIE AFGESLOTEN | Gebruiker: %s", user)
        logging.info("=" * 70)

        response = jsonify({"success": True, "message": "Applicatie sluit af"})

        def shutdown_server() -> None:
            import time

            time.sleep(1)
            logging.info("Flask server wordt beeindigd...")
            os._exit(0)

        import threading

        shutdown_thread = threading.Thread(target=shutdown_server, daemon=True)
        shutdown_thread.start()

        return response, 200
    except Exception as exc:
        logging.error("Fout bij afsluiten applicatie: %s", str(exc))
        return jsonify({"success": False, "message": f"Fout: {str(exc)}"}), 500


@app.route("/settings")
def settings():
    if MAIN_APP_URL:
        return redirect(f"{MAIN_APP_URL}/settings")
    return jsonify({"success": False, "message": "Instellingen zijn alleen beschikbaar via de hoofdapp."}), 403


if __name__ == "__main__":
    if not os.path.exists(LOG_DIRECTORY):
        try:
            os.makedirs(LOG_DIRECTORY)
        except Exception as exc:
            print(f"FOUT: Kan log directory niet aanmaken: {LOG_DIRECTORY}")
            print(f"Details: {str(exc)}")
            exit(1)

    log_file_path = os.path.join(LOG_DIRECTORY, "transactietoevoegen_webapp_log.txt")
    logging.basicConfig(
        filename=log_file_path,
        level=getattr(logging, str(LOG_LEVEL).upper(), logging.INFO),
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    logging.info("=" * 70)
    logging.info("TRANSACTIE TOEVOEGEN START")
    logging.info("Bank bestand: %s", config.get("bank_excel_path", ""))
    logging.info("=" * 70)

    port = int(os.getenv("DEBUTADE_APP_PORT", "5004"))
    app.run(debug=False, host="127.0.0.1", port=port, use_reloader=False)
