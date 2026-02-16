"""
Verbeterde tag-recommender met TF-IDF + SVM/SGD classifier en class-balancing.
Ondersteunt Nederlandse stopwoorden, bedrag-binning, en confidence-threshold filtering.
"""
import logging
import math
import os
import re
from collections import Counter, defaultdict
from typing import Dict, List, Tuple

import numpy as np
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer, ENGLISH_STOP_WORDS
from sklearn.svm import LinearSVC
from sklearn.pipeline import make_pipeline
from sklearn.utils.class_weight import compute_class_weight

TOKEN_RE = re.compile(r"[A-Za-z0-9]+")

# Nederlandse stopwoorden (basisset)
DUTCH_STOP_WORDS = ENGLISH_STOP_WORDS | {
    "de", "het", "een", "en", "of", "voor", "met", "door", "in", "op", "aan",
    "naar", "van", "tot", "als", "bij", "uit", "is", "zijn", "ben", "wordt",
    "om", "te", "geen", "dit", "dat", "deze", "die", "mijn", "jouw", "ons",
    "hun", "wat", "wie", "welk", "waar", "wanneer", "hoe", "waarom",
}


def _create_bedrag_bin(bedrag: float) -> str:
    """Verdeel bedrag in bins voor betere feature engineering."""
    if bedrag < 0:
        return "BEDRAG_NEG"
    elif bedrag == 0:
        return "BEDRAG_ZERO"
    elif bedrag < 10:
        return "BEDRAG_TINY"
    elif bedrag < 50:
        return "BEDRAG_SMALL"
    elif bedrag < 200:
        return "BEDRAG_MEDIUM"
    elif bedrag < 1000:
        return "BEDRAG_LARGE"
    else:
        return "BEDRAG_XLARGE"


class TagRecommender:
    """Houdt een lichtgewicht vocabulaire per tag bij en kan suggesties genereren."""

    def __init__(self, training_path: str, allowed_tags: List[str] | None = None, additional_data_path: str | None = None, confidence_threshold: float = 0.15):
        self.training_path = training_path
        self.additional_data_path = additional_data_path  # Bijv. werkbestand met al ingevulde tags
        self.allowed_tags = set(allowed_tags or [])
        self.tag_token_freq: defaultdict[str, Counter[str]] = defaultdict(Counter)
        self.token_doc_freq: Counter[str] = Counter()
        self.tag_totals: Counter[str] = Counter()
        self.total_docs = 0
        self.last_loaded_mtime: float | None = None
        self.last_additional_mtime: float | None = None
        self.model = None
        self.confidence_threshold = confidence_threshold  # Min. confidence voor ML-suggestie
        self.class_weights = None  # Voor class-balancing

    @staticmethod
    def _tokenize(text: str) -> List[str]:
        """Tokenizeer tekst en voeg samengestelde woorden toe."""
        # Basis tokenisatie
        basic_tokens = [match.group(0).lower() for match in TOKEN_RE.finditer(text or "")]
        
        # Filter zeer korte tokens (< 2 tekens)
        basic_tokens = [t for t in basic_tokens if len(t) >= 2]
        
        # Voeg aanvullende tokens toe voor betere matching
        extra_tokens = []
        for token in basic_tokens:
            if "jeugd" in token:
                extra_tokens.append("jeugd")
            if "volwassenen" in token or "volwassen" in token:
                extra_tokens.append("volwassenen")
        
        return basic_tokens + extra_tokens

    def _reset(self) -> None:
        self.tag_token_freq.clear()
        self.token_doc_freq.clear()
        self.tag_totals.clear()
        self.total_docs = 0

    def _find_columns(self, header: List[str]) -> Tuple[int | None, List[int]]:
        """Zoek de kolommen voor tag en tekstvelden."""
        normalized = [str(col).strip().lower() for col in header]
        lookup = {name: idx for idx, name in enumerate(normalized)}

        tag_col = None
        for candidate in ("tag", "tags", "categorie", "category"):
            if candidate in lookup:
                tag_col = lookup[candidate]
                break

        text_cols = []
        for candidate in (
            "naam / omschrijving",
            "naam/omschrijving",
            "mededeling",
            "mededelingen",
            "omschrijving",
            "rekening",
            "tegenrekening",
            "mutatiesoort",
            "memo",
            "code",
            "description",
        ):
            if candidate in lookup:
                text_cols.append(lookup[candidate])

        if not text_cols:
            # Geen duidelijke tekstkolommen gevonden: gebruik alle kolommen behalve de tagkolom
            text_cols = [idx for idx in range(len(header)) if idx != tag_col]

        return tag_col, text_cols

    def _collect_dataset(self, path: str) -> List[tuple[str, str]]:
        """Lees een Excelbestand en verzamel (text, tag) voorbeelden met verbeterde features."""
        samples: List[tuple[str, str]] = []
        wb = None
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            for sheet in wb.worksheets:
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                if not first_row:
                    continue
                header = [str(val).strip() if val is not None else "" for val in first_row]
                normalized = [str(col).strip().lower() for col in header]
                lookup = {name: idx for idx, name in enumerate(normalized)}

                # Zoek tag kolom
                tag_col = None
                for candidate in ("tag", "tags", "categorie", "category"):
                    if candidate in lookup:
                        tag_col = lookup[candidate]
                        break
                if tag_col is None:
                    continue

                # Zoek tekstkolommen met prioriteit
                text_cols = []
                priority_candidates = [
                    ("mededeling", "mededelingen"),  # Highest priority
                    ("naam / omschrijving", "naam/omschrijving", "omschrijving"),
                    ("mutatiesoort", "code"),
                    ("rekening", "tegenrekening"),
                ]
                for candidate_group in priority_candidates:
                    for candidate in candidate_group:
                        if candidate in lookup and lookup[candidate] not in text_cols:
                            text_cols.append(lookup[candidate])
                
                if not text_cols:
                    text_cols = [idx for idx in range(len(header)) if idx != tag_col]

                # Zoek bedrag kolom (optioneel)
                amount_col = None
                for candidate in ("bedrag (eur)", "bedrag", "amount"):
                    if candidate in lookup:
                        amount_col = lookup[candidate]
                        break

                # Zoek af/bij kolom (sign indicator)
                sign_col = None
                for candidate in ("af bij", "af/bij", "afbij", "sign"):
                    if candidate in lookup:
                        sign_col = lookup[candidate]
                        break

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) <= tag_col:
                        continue
                    tag_val = str(row[tag_col] or "").strip()
                    if not tag_val:
                        continue
                    if self.allowed_tags and tag_val not in self.allowed_tags:
                        continue

                    # Bouw feature string met gewichten
                    parts: List[str] = []
                    
                    # Hoogste prioriteit: mededelingen/omschrijving (meest informatief)
                    for idx in text_cols[:2]:  # Top 2 kolommen krijgen meeste gewicht
                        if idx < len(row) and row[idx] not in (None, ""):
                            val_str = str(row[idx]).strip()
                            if val_str and len(val_str) > 1:
                                # Repeat voor meer gewicht in TF-IDF
                                parts.extend([val_str, val_str])
                    
                    # Lagere prioriteit: overige kolommen
                    for idx in text_cols[2:]:
                        if idx < len(row) and row[idx] not in (None, ""):
                            val_str = str(row[idx]).strip()
                            if val_str and len(val_str) > 1:
                                parts.append(val_str)

                    # Voeg bedrag-bin toe als feature
                    if amount_col is not None and amount_col < len(row) and row[amount_col] not in (None, ""):
                        try:
                            amount_val = float(str(row[amount_col]).replace(",", "."))
                            bedrag_bin = _create_bedrag_bin(amount_val)
                            parts.append(bedrag_bin)
                        except (ValueError, TypeError):
                            pass

                    # Voeg af/bij indicator toe
                    if sign_col is not None and sign_col < len(row) and row[sign_col] not in (None, ""):
                        sign_str = str(row[sign_col]).strip().lower()
                        if sign_str:
                            parts.append(f"SIGN_{sign_str}")

                    if not parts:
                        continue

                    combined = " ".join(parts)
                    samples.append((combined, tag_val))
        except Exception as exc:  # noqa: BLE001
            logging.error("Fout bij laden dataset uit %s: %s", path, exc)
        finally:
            if wb:
                wb.close()
        return samples

    def _process_heuristic_sample(self, text: str, tag: str) -> None:
        """Verwerk een sample voor heuristische benadering."""
        tokens = self._tokenize(text)
        for token in tokens:
            self.tag_token_freq[tag][token] += 1
            self.token_doc_freq[token] += 1
        self.tag_totals[tag] += 1
        self.total_docs += 1

    def load(self) -> bool:
        """Train het ML-model op trainingsdata + reeds getagde werkdata met class-balancing."""
        if not self.training_path or not os.path.exists(self.training_path):
            logging.warning("Trainingsbestand niet gevonden: %s", self.training_path)
            return False

        # Hertrain alleen als bronbestanden gewijzigd zijn
        mtimes = [os.path.getmtime(self.training_path)]
        if self.additional_data_path and os.path.exists(self.additional_data_path):
            mtimes.append(os.path.getmtime(self.additional_data_path))
        latest_mtime = max(mtimes)
        if self.last_loaded_mtime and latest_mtime <= self.last_loaded_mtime:
            return True

        self._reset()

        # Verzamel training samples
        samples = self._collect_dataset(self.training_path)
        if self.additional_data_path and os.path.exists(self.additional_data_path):
            samples += self._collect_dataset(self.additional_data_path)

        if not samples:
            logging.warning("Geen trainingsdata gevonden om model te trainen")
            return False

        texts, labels = zip(*samples)
        labels_list = list(labels)

        # Controleer aantal unieke klassen
        unique_classes = set(labels_list)
        if len(unique_classes) < 2:
            logging.warning(
                "Onvoldoende trainingsklassen (%d) voor ML model; gebruik heuristische benadering",
                len(unique_classes)
            )
            # Bouw heuristische tag-vocabulaire
            for text, label in samples:
                self._process_heuristic_sample(text, label)
            self.model = None  # Markeer dat heuristics gebruikt worden
            self.last_loaded_mtime = latest_mtime
            return True

        # Bereken class weights voor balancing
        try:
            unique_labels = np.unique(labels_list)
            self.class_weights = compute_class_weight('balanced', classes=unique_labels, y=np.array(labels_list))
            class_weight_dict = dict(zip(unique_labels, self.class_weights))
            logging.info("Class weights berekend: %s", class_weight_dict)
        except Exception as exc:
            logging.warning("Kon class weights niet berekenen: %s", exc)
            class_weight_dict = 'balanced'

        # ML pipeline: TF-IDF met Nederlandse stopwoorden + LinearSVC (beter dan Logistic Regression voor tekst)
        model = make_pipeline(
            TfidfVectorizer(
                ngram_range=(1, 2),
                min_df=1,
                max_df=0.95,  # Exclude zeer frequente woorden
                stop_words=list(DUTCH_STOP_WORDS),
                lowercase=True,
                strip_accents='unicode',
                norm='l2',
                use_idf=True,
            ),
            LinearSVC(
                C=1.0,
                max_iter=2000,
                class_weight=class_weight_dict,
                random_state=42,
                dual=False,
                loss='squared_hinge',
                verbose=0,
            )
        )

        try:
            model.fit(texts, labels_list)
            self.model = model
            self.last_loaded_mtime = latest_mtime
            logging.info("ML model (LinearSVC) getraind met %d voorbeelden over %d klassen", len(samples), len(unique_classes))
            return True
        except Exception as exc:
            logging.error("ML model training mislukt: %s; valt terug op heuristics", exc)
            # Fallback: bouw heuristische tag-vocabulaire
            for text, label in samples:
                self._process_heuristic_sample(text, label)
            self.model = None
            self.last_loaded_mtime = latest_mtime
            return True

    def recommend(self, transaction: Dict[str, str], top_k: int = 3) -> List[Dict[str, float | str]]:
        """Geef een lijst met tags en scores terug op basis van het ML-model of heuristics.
        
        Gebruikt confidence-threshold: alleen ML-suggesties boven de drempel worden gegeven.
        """
        if not self.load():
            return []

        parts: List[str] = []
        
        # Prioritized field collection
        high_priority_fields = ("mededelingen", "omschrijving", "naam")
        medium_priority_fields = ("mutatiesoort", "code")
        low_priority_fields = ("rekening", "tegenrekening", "memo")
        
        # High priority: repeat 2x voor extra gewicht
        for key in high_priority_fields:
            val = transaction.get(key)
            if val:
                val_str = str(val).strip()
                if len(val_str) > 1:
                    parts.extend([val_str, val_str])
        
        # Medium priority: repeat 1x
        for key in medium_priority_fields:
            val = transaction.get(key)
            if val:
                val_str = str(val).strip()
                if len(val_str) > 1:
                    parts.append(val_str)
        
        # Low priority: single mention
        for key in low_priority_fields:
            val = transaction.get(key)
            if val:
                val_str = str(val).strip()
                if len(val_str) > 1:
                    parts.append(val_str)

        # Voeg bedrag-bin toe als feature
        bedrag_str = str(transaction.get("bedrag", "")).strip()
        try:
            bedrag = float(bedrag_str.replace(",", ".")) if bedrag_str else None
        except (ValueError, AttributeError):
            bedrag = None
        if bedrag is not None:
            bedrag_bin = _create_bedrag_bin(bedrag)
            parts.append(bedrag_bin)

        # Voeg af/bij indicator toe
        af_bij_str = str(transaction.get("af_bij", "")).strip().lower()
        if af_bij_str:
            parts.append(f"SIGN_{af_bij_str}")

        if not parts:
            return []

        text = " ".join(parts)

        # Probeer ML-model te gebruiken
        if hasattr(self, "model") and self.model is not None:
            try:
                # LinearSVC voorspelt discrete klassen; probeer decision_function scores
                if hasattr(self.model.named_steps['linearsvc'], 'decision_function'):
                    scores = self.model.named_steps['linearsvc'].decision_function([
                        self.model.named_steps['tfidfvectorizer'].transform([text]).toarray()[0]
                    ])[0]
                    
                    # Normaliseer scores naar [0, 1] (sigmoid-achtig)
                    from scipy.special import expit
                    proba = expit(scores)  # Logistic sigmoid
                else:
                    # Fallback: gebruik predict_proba als beschikbaar
                    proba = self.model.predict_proba([text])[0]
                
                classes = self.model.classes_
                paired = sorted(zip(classes, proba), key=lambda p: p[1], reverse=True)
                
                # Filter op confidence threshold
                confident_results = []
                for tag, score in paired:
                    if score >= self.confidence_threshold:
                        confident_results.append({"tag": tag, "score": round(float(score), 4)})
                    if len(confident_results) >= top_k:
                        break
                
                if confident_results:
                    return confident_results
                
                # Fallback: geen scores boven threshold, retourneer top-1 wel (beter dan niets)
                if paired:
                    logging.debug("Laag vertrouwen (%f) voor suggestie, maar retourneer top-1: %s (score: %f)", 
                                 paired[0][1], paired[0][0], paired[0][1])
                    return [{"tag": paired[0][0], "score": round(float(paired[0][1]), 4)}]
                
            except Exception as exc:  # noqa: BLE001
                logging.debug("Fout bij ML aanbeveling (val terug op heuristics): %s", exc)

        # Fallback: heuristische benadering
        tokens = self._tokenize(text)
        tag_scores: Dict[str, float] = {}

        for tag in self.tag_token_freq:
            score = 0.0
            for token in tokens:
                if token in self.tag_token_freq[tag]:
                    # TF-IDF-achtige scoring
                    tf = self.tag_token_freq[tag][token]
                    idf = math.log(self.total_docs / max(self.token_doc_freq[token], 1)) if self.total_docs > 0 else 0
                    score += tf * idf
            if score > 0:
                tag_scores[tag] = score

        # Filter op allowed_tags
        if self.allowed_tags:
            tag_scores = {tag: score for tag, score in tag_scores.items() if tag in self.allowed_tags}

        if not tag_scores:
            return []

        sorted_tags = sorted(tag_scores.items(), key=lambda p: p[1], reverse=True)
        return [
            {"tag": tag, "score": round(float(score), 4)}
            for tag, score in sorted_tags[:top_k]
        ]
