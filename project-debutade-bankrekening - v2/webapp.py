"""
Bankrekening Debutade - Web Applicatie
=======================================

Een moderne web-gebaseerde applicatie voor het beheren van bankrekeningtransacties.
Dit is de Flask web app versie van de originele Tkinter applicatie.

Functionaliteiten:
- Invoeren van financiële gegevens via een webinterface
- Validatie van invoer (datums en bedragen)
- Automatische opslag in Excel-bestand
- Logging van gebeurtenissen
- Overzicht van recente transacties
- Berekening van totaal banksaldo

Versie: 2.0 (Web App)
Datum: 2026-01-03
Auteur: Eric G.
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, send_from_directory
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import json
import logging
import shutil
import locale
import getpass
import sys

try:
    from tag_recommender import TagRecommender
except ModuleNotFoundError:
    import sys as _sys
    import os as _os
    _sys.path.append(_os.path.dirname(_os.path.abspath(__file__)))
    from tag_recommender import TagRecommender

# Fix encoding voor Windows console
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Vereiste kolom headers voor het Excel bestand
REQUIRED_HEADERS = [
    "Datum",
    "Naam / Omschrijving",
    "Rekening",
    "Tegenrekening",
    "Code",
    "Af Bij",
    "Bedrag (EUR)",
    "Mutatiesoort",
    "Mededelingen",
    "Saldo na mutatie",
    "",
    "Tag"
]

# Vereiste tabs (sheets) in het Excel bestand
REQUIRED_SHEETS = [
    "Bankrekening",
    "Spaarrekening 1",
    "Spaarrekening 2"
]

app = Flask(__name__)
app.static_folder = 'static'
# Zorg dat gewijzigde templates direct opnieuw geladen worden (ontwikkelmodus)
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Laad configuratie
def load_config(config_path, section_key="bankrekening"):
    """Laad de configuratie uit een JSON bestand"""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Configuratiebestand niet gevonden: {config_path}")

    with open(config_path, "r", encoding="utf-8") as config_file:
        root_config = json.load(config_file)

    if section_key not in root_config:
        raise KeyError(f"Configuratiesectie ontbreekt: {section_key}")

    config = root_config[section_key]
    shared = root_config.get("shared", {})
    for key in ("backup_directory", "log_directory", "resources", "log_level", "tags"):
        if key in shared:
            config[key] = shared[key]
    if shared.get("grootboek_directory") and config.get("excel_file_name"):
        config["excel_file_path"] = os.path.join(
            shared["grootboek_directory"],
            config["excel_file_name"],
        )

    required_keys = [
        "excel_file_name",
        "excel_sheet_name",
        "required_sheets",
        "backup_directory",
        "log_directory",
        "resources",
        "tags",
        "log_level",
    ]

    for key in required_keys:
        if key not in config:
            raise KeyError(f"Configuratiesleutel ontbreekt: {key}")

    return config

# Bepaal het directory waar het script zich bevindt
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Standaard configuratie pad (kan worden aangepast via omgevingsvariabele)
CONFIG_PATH = os.getenv(
    "DEBUTADE_CONFIG",
    os.path.abspath(os.path.join(SCRIPT_DIR, "..", "config.json")),
)

try:
    config = load_config(CONFIG_PATH)
except (FileNotFoundError, KeyError) as e:
    print(f"WAARSCHUWING: {e}")
    # Gebruik standaard configuratie voor ontwikkeling
    config = {
        "excel_file_directory": r"C:\Users\ericg\OneDrive\Documents\Code",
        "excel_file_name": "records.xlsx",
        "resources": r"C:\Users\ericg\OneDrive\Documents\Code\resources",
        "backup_directory": r"C:\Users\ericg\OneDrive\Documents\Code\backups",
        "log_directory": r"C:\Users\ericg\OneDrive\Documents\Code\logs",
        "excel_sheet_name": "Bankrekening",
        "tags": ["Algemeen", "Evenement", "Materiaal", "Training", "Overig"],
        "log_level": "INFO"
    }

# Configuratie variabelen
EXCEL_FILE_PATH = config["excel_file_path"]
EXCEL_FILE_DIRECTORY = os.path.dirname(EXCEL_FILE_PATH)
EXCEL_FILE_NAME = os.path.basename(EXCEL_FILE_PATH)
BACKUP_DIRECTORY = config["backup_directory"]
LOG_DIRECTORY = config["log_directory"]
EXCEL_SHEET_NAME = config["excel_sheet_name"]
TAGS = config["tags"]
LOG_LEVEL = config["log_level"]
REQUIRED_SHEETS = config.get("required_sheets", REQUIRED_SHEETS)
TRAINING_FILE_PATH = os.path.join(SCRIPT_DIR, "static", "category_test_set.xlsx")
MAIN_APP_URL = os.getenv("MAIN_APP_URL", "").strip()


def settings_locked_response():
    message = "Instellingen zijn alleen beschikbaar via de hoofdapp."
    if MAIN_APP_URL:
        return jsonify({
            "success": False,
            "message": message,
            "redirect": f"{MAIN_APP_URL}/settings",
        }), 403
    return jsonify({"success": False, "message": message}), 403

# Initialiseer TagRecommender met trainingsdata en werkbestand als aanvullende data
tag_recommender = TagRecommender(TRAINING_FILE_PATH, allowed_tags=TAGS, additional_data_path=EXCEL_FILE_PATH)
tag_recommender.load()

# Fallback: bepaal tag op basis van meest gebruikte tag voor dezelfde tegenrekening
def suggest_tag_by_tegenrekening(tegenrekening: str) -> str | None:
    wb = None
    try:
        tegen = str(tegenrekening or "").strip().upper()
        if not tegen or not os.path.exists(EXCEL_FILE_PATH):
            return None

        wb = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
        tag_counts: dict[str, int] = {}
        for sheet_name in REQUIRED_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_tegen = str((row[3] if len(row) > 3 else "") or "").strip().upper()
                tag_val = str((row[11] if len(row) > 11 else "") or "").strip()
                if row_tegen and tag_val and row_tegen == tegen:
                    tag_counts[tag_val] = tag_counts.get(tag_val, 0) + 1

        if not tag_counts:
            return None
        # Kies de tag met de hoogste frequentie
        return max(tag_counts.items(), key=lambda kv: kv[1])[0]
    except Exception as e:  # noqa: BLE001
        logging.error(f"Fout bij fallback suggestie op basis van tegenrekening: {str(e)}")
        return None
    finally:
        if wb:
            wb.close()

# Valideer alle bestandspaden bij startup
def validate_config():
    """Valideer configuratie - start wel maar waarschuw als Excel pad leeg is"""
    
    logging.info("=" * 70)
    logging.info("CONFIGURATIE VALIDATIE")
    logging.info(f"Excel pad uit config: '{EXCEL_FILE_PATH}'")
    
    # Controleer of Excel pad leeg is
    if not EXCEL_FILE_PATH or EXCEL_FILE_PATH.strip() == "":
        logging.warning("Excel bestandspad is LEEG")
        logging.warning("Gebruiker MOET eerst via Instellingen in de hoofdapp een bestand selecteren!")
        return True  # App start wel, maar gebruiker moet eerst Excel bestand kiezen
    else:
        logging.info(f"Bestand bestaat: {os.path.exists(EXCEL_FILE_PATH)}")
        
        # Controleer of Excel bestand bestaat
        if not os.path.exists(EXCEL_FILE_PATH):
            logging.warning(f"Excel bestand niet gevonden: {EXCEL_FILE_PATH}")
            # Extra debugging
            dir_path = os.path.dirname(EXCEL_FILE_PATH)
            if os.path.exists(dir_path):
                logging.info(f"Directory bestaat wel: {dir_path}")
                try:
                    files = os.listdir(dir_path)
                    xlsx_files = [f for f in files if f.endswith('.xlsx')]
                    logging.info(f".xlsx bestanden in directory: {xlsx_files}")
                except:
                    pass
            else:
                logging.warning(f"Directory bestaat niet: {dir_path}")
            return True  # App start wel
        else:
            logging.info(f"Excel bestand gevonden: {EXCEL_FILE_PATH}")
    
    # Controleer of directories bestaan, anders aanmaken
    for dir_name, dir_path in [("Backup", BACKUP_DIRECTORY), ("Log", LOG_DIRECTORY), ("Excel", EXCEL_FILE_DIRECTORY)]:
        if not os.path.exists(dir_path):
            try:
                os.makedirs(dir_path)
                logging.info(f"{dir_name} directory aangemaakt: {dir_path}")
            except Exception as e:
                logging.error(f"Kan {dir_name} directory niet aanmaken: {dir_path}")
        else:
            logging.info(f"{dir_name} directory bestaat: {dir_path}")
    
    # Controleer of log file schrijfbaar is
    if os.path.exists(LOG_DIRECTORY):
        try:
            test_log = os.path.join(LOG_DIRECTORY, ".write_test")
            with open(test_log, "w") as f:
                f.write("test")
            os.remove(test_log)
            logging.info("Log directory is schrijfbaar")
        except Exception as e:
            logging.error(f"Log directory is niet schrijfbaar: {str(e)}")
    
    logging.info("Applicatie start!")
    logging.info("=" * 70)
    return True

# Stel Nederlandse locale in (optioneel, kan problemen geven op sommige systemen)
try:
    locale.setlocale(locale.LC_TIME, "nl_NL")
except:
    pass

# Maak backup bij opstarten
def create_backup():
    """Maak een backup van het Excel bestand"""
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(BACKUP_DIRECTORY, 
                f"{EXCEL_FILE_NAME}_backup_{timestamp}.xlsx")
            shutil.copy(EXCEL_FILE_PATH, backup_path)
            logging.info(f"Backup gemaakt: {backup_path}")
            return True
    except Exception as e:
        logging.error(f"Fout bij maken backup: {str(e)}")
        return False

def calculate_total_amount():
    """Bereken het totale saldo in de kas"""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return 0
        
        wb = load_workbook(EXCEL_FILE_PATH)
        if EXCEL_SHEET_NAME in wb.sheetnames:
            sheet = wb[EXCEL_SHEET_NAME]
            total = 0
            # Kolom F = Af/Bij (kolom 6), Kolom G = Bedrag (kolom 7)
            for row in sheet.iter_rows(min_row=2, min_col=6, max_col=7, values_only=True):
                af_bij, amount = row
                if isinstance(amount, (int, float)):
                    if af_bij == "Af":
                        total -= amount
                    elif af_bij == "Bij":
                        total += amount
            return round(total, 2)
        return 0
    except Exception as e:
        logging.error(f"Fout bij berekenen totaal: {str(e)}")
        return 0

def get_recent_transactions(limit=10):
    """Haal de meest recente transacties op"""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return []
        
        wb = load_workbook(EXCEL_FILE_PATH)
        if EXCEL_SHEET_NAME not in wb.sheetnames:
            return []
        
        sheet = wb[EXCEL_SHEET_NAME]
        transactions = []
        
        # Start bij rij 2 (rij 1 is header)
        for row in sheet.iter_rows(min_row=2, max_row=min(limit+1, sheet.max_row), 
                                   values_only=True):
            if row[0]:  # Als datum bestaat
                transactions.append({
                    'datum': row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0]),
                    'mededelingen': (row[8] if len(row) > 8 else None) or row[1] or '',
                    'af_bij': row[5] or '',
                    'bedrag': f"{row[6]:.2f}" if isinstance(row[6], (int, float)) else '0.00',
                    'tag': row[11] or '',
                    'saldo': f"€ {row[9]:.2f}" if isinstance(row[9], (int, float)) else '€ 0.00'
                })
        
        return transactions
    except Exception as e:
        logging.error(f"Fout bij ophalen transacties: {str(e)}")
        return []

def get_all_transactions():
    """Haal alle transacties op uit het Excel bestand"""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return []
        
        wb = load_workbook(EXCEL_FILE_PATH)
        if EXCEL_SHEET_NAME not in wb.sheetnames:
            return []
        
        sheet = wb[EXCEL_SHEET_NAME]
        transactions = []
        
        # Start bij rij 2 (rij 1 is header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Als datum bestaat
                transactions.append({
                    'datum': row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0]),
                    'mededelingen': (row[8] if len(row) > 8 else None) or row[1] or '',
                    'af_bij': row[5] or '',
                    'bedrag': f"{row[6]:.2f}" if isinstance(row[6], (int, float)) else '0.00',
                    'rekening': row[2] or '',
                    'tag': row[11] or ''
                })
        
        return transactions
    except Exception as e:
        logging.error(f"Fout bij ophalen alle transacties: {str(e)}")
        return []

def get_untagged_transactions():
    """Haal alle transacties op zonder ingevulde Tag (leeg of whitespace) uit alle vereiste tabs."""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return []

        wb = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
        transactions = []

        for sheet_name in REQUIRED_SHEETS:
            if sheet_name not in wb.sheetnames:
                # Als een vereiste sheet ontbreekt, sla over; validatie elders bewaakt structuur
                continue
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                tag_value = (row[11] if len(row) > 11 else '') or ''
                if str(tag_value).strip() == '':
                    transactions.append({
                        'sheet_name': sheet_name,
                        'row_index': row_idx,
                        'datum': row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else (row[0] or ''),
                        'mededelingen': (row[8] if len(row) > 8 else None) or row[1] or '',
                        'af_bij': row[5] or '',
                        'bedrag': f"{row[6]:.2f}" if isinstance(row[6], (int, float)) else '0.00',
                        'rekening': row[2] or ''
                    })

        return transactions
    except Exception as e:
        logging.error(f"Fout bij ophalen ongetagde transacties: {str(e)}")
        return []

def get_all_transactions_all_sheets():
    """Haal alle transacties uit alle vereiste tabs, inclusief bestaande Tag."""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return []
        wb = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
        transactions = []

        for sheet_name in REQUIRED_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row and row[0]:
                    transactions.append({
                        'sheet_name': sheet_name,
                        'row_index': row_idx,
                        'datum': row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else (row[0] or ''),
                        'mededelingen': (row[8] if len(row) > 8 else None) or row[1] or '',
                        'af_bij': row[5] or '',
                        'bedrag': f"{row[6]:.2f}" if isinstance(row[6], (int, float)) else '0.00',
                        'rekening': row[2] or '',
                        'tag': (row[11] if len(row) > 11 else '') or ''
                    })

        return transactions
    except Exception as e:
        logging.error(f"Fout bij ophalen alle transacties (alle tabs): {str(e)}")
        return []


def get_transaction_from_sheet(sheet_name, row_index):
    """Lees een enkele rij uit de opgegeven sheet voor AI-suggesties."""
    wb = None
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return None, "Excel bestand niet gevonden"
        wb = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None, "Sheet niet gevonden in Excel bestand"

        sheet = wb[sheet_name]
        row = next(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True), None)
        if not row:
            return None, "Rij niet gevonden in sheet"

        transaction = {
            'datum': row[0] if len(row) > 0 else '',
            'omschrijving': row[1] if len(row) > 1 else '',
            'mededelingen': (row[8] if len(row) > 8 else None) or (row[1] if len(row) > 1 else ''),
            'rekening': row[2] if len(row) > 2 else '',
            'tegenrekening': row[3] if len(row) > 3 else '',
            'code': row[4] if len(row) > 4 else '',
            'af_bij': row[5] if len(row) > 5 else '',
            'bedrag': row[6] if len(row) > 6 else '',
            'mutatiesoort': row[7] if len(row) > 7 else '',
        }
        return transaction, None
    except Exception as e:  # noqa: BLE001
        logging.error(f"Fout bij lezen van transactie voor AI-suggestie: {str(e)}")
        return None, f"Fout bij lezen van transactie: {str(e)}"
    finally:
        if wb:
            wb.close()

def get_sheet_stats():
    """Geef per vereiste tab het aantal rijen en aantal ongetagde rijen terug."""
    stats = []
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return stats
        wb = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
        for sheet_name in REQUIRED_SHEETS:
            total_rows = 0
            untagged_rows = 0
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row and any(cell is not None and str(cell).strip() != '' for cell in row):
                        total_rows += 1
                        tag_value = (row[11] if len(row) > 11 else '') or ''
                        if str(tag_value).strip() == '':
                            untagged_rows += 1
            stats.append({'sheet_name': sheet_name, 'total': total_rows, 'untagged': untagged_rows})
        return stats
    except Exception as e:
        logging.error(f"Fout bij ophalen sheet statistieken: {str(e)}")
        return stats
@app.route('/favicon.ico')
def favicon():
    """Serve the favicon"""
    return send_from_directory(app.static_folder, 'icon.ico', mimetype='image/vnd.microsoft.icon')

@app.route('/')
def index():
    """Hoofdpagina met invoerformulier"""
    total_amount = calculate_total_amount()
    untagged_transactions = get_untagged_transactions()
    all_transactions = get_all_transactions_all_sheets()
    sheet_stats = get_sheet_stats()
    today = datetime.now().strftime('%Y-%m-%d')
    current_date_display = datetime.now().strftime('%d-%m-%Y')
    current_user = getpass.getuser()
    
    return render_template('index.html', 
                         tags=TAGS,
                         total_amount=total_amount,
                         untagged_transactions=untagged_transactions,
                         all_transactions=all_transactions,
                         sheet_stats=sheet_stats,
                         today=today,
                         current_date=current_date_display,
                         current_user=current_user,
                         main_app_url=MAIN_APP_URL)


@app.route('/recommend_tag', methods=['POST'])
def recommend_tag():
    """Geef een tag-suggestie op basis van de trainingsset."""
    try:
        if not EXCEL_FILE_PATH or not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({'success': False, 'message': 'Excel bestand niet beschikbaar'}), 400

        data = request.get_json() or {}
        sheet_name = str(data.get('sheet_name', '')).strip()
        row_index = int(str(data.get('row_index', '0')).strip() or '0')

        if sheet_name == '' or sheet_name not in REQUIRED_SHEETS:
            return jsonify({'success': False, 'message': 'Ongeldige sheet-naam'}), 400
        if row_index < 2:
            return jsonify({'success': False, 'message': 'Ongeldige rij-index'}), 400

        transaction, error_message = get_transaction_from_sheet(sheet_name, row_index)
        if error_message:
            return jsonify({'success': False, 'message': error_message}), 400

        suggestions = tag_recommender.recommend(transaction, top_k=3) if tag_recommender else []
        if not suggestions:
            # Fallback: probeer op basis van tegenrekening
            fallback_tag = suggest_tag_by_tegenrekening(transaction.get('tegenrekening'))
            if fallback_tag:
                suggestions = [{'tag': fallback_tag, 'score': 1.0}]
            else:
                return jsonify({'success': False, 'message': 'Geen suggesties beschikbaar op basis van trainingsset of tegenrekening.'}), 404

        return jsonify({'success': True, 'top_tag': suggestions[0]['tag'], 'suggestions': suggestions})
    except Exception as e:  # noqa: BLE001
        logging.error(f"Fout bij genereren tag-suggestie: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500


@app.route('/bulk_recommend_tags', methods=['POST'])
def bulk_recommend_tags():
    """Pas AI suggesties toe op alle transacties zonder tag, behalve "Beginsaldo" transacties."""
    try:
        if not EXCEL_FILE_PATH or not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({'success': False, 'message': 'Excel bestand niet beschikbaar'}), 400

        wb = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
        results = []
        
        for sheet_name in REQUIRED_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
                
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 12:
                    continue
                
                # Controleer of tag leeg is (kolom 12, index 11)
                tag_val = str(row[11] or "").strip()
                if tag_val:
                    continue  # Skip rijen die al een tag hebben
                
                # Controleer of deze rij "Beginsaldo" bevat in ALLE tekstkolommen
                # Kolom 1 (index 0) = Naam, Kolom 8 (index 7) = Mutatiesoort, Kolom 9 (index 8) = Mededelingen
                naam = str(row[0] or "").strip().lower()
                mutatiesoort = str(row[7] or "").strip().lower() if len(row) > 7 else ""
                mededelingen = str(row[8] or "").strip().lower() if len(row) > 8 else ""
                
                if "beginsaldo" in naam or "beginsaldo" in mutatiesoort or "beginsaldo" in mededelingen:
                    continue  # Skip beginsaldo transacties
                
                # Bouw transaction object
                transaction = {
                    'datum': str(row[0] or ""),
                    'naam': str(row[1] or ""),
                    'rekening': str(row[2] or ""),
                    'tegenrekening': str(row[3] or ""),
                    'code': str(row[4] or ""),
                    'af_bij': str(row[5] or ""),
                    'bedrag': str(row[6] or ""),
                    'mutatiesoort': str(row[7] or ""),
                    'mededelingen': mededelingen,
                    'omschrijving': str(row[1] or "")
                }
                
                # Vraag AI suggestie op (top 3)
                suggestions = tag_recommender.recommend(transaction, top_k=3) if tag_recommender else []

                if not suggestions:
                    # Fallback: probeer op basis van tegenrekening
                    fallback_tag = suggest_tag_by_tegenrekening(transaction.get('tegenrekening'))
                    if fallback_tag:
                        suggestions = [{'tag': fallback_tag, 'score': 1.0}]

                if suggestions:
                    results.append({
                        'success': True,
                        'sheet_name': sheet_name,
                        'row_index': row_idx,
                        'tag': suggestions[0]['tag'],
                        'suggestions': suggestions
                    })
                else:
                    results.append({
                        'success': False,
                        'sheet_name': sheet_name,
                        'row_index': row_idx,
                        'message': 'Geen suggestie beschikbaar',
                        'suggestions': []
                    })
        
        wb.close()
        return jsonify({'success': True, 'results': results, 'count': len([r for r in results if r['success']])})
    
    except Exception as e:  # noqa: BLE001
        logging.error(f"Fout bij bulk AI suggesties: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500


@app.route('/update_tag', methods=['POST'])
def update_tag():
    """Werk de Tag bij voor een specifieke rij in een opgegeven sheet."""
    try:
        if not EXCEL_FILE_PATH or not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({'success': False, 'message': 'Excel bestand niet beschikbaar'}), 400

        data = request.get_json() or {}
        sheet_name = str(data.get('sheet_name', '')).strip()
        row_index = int(str(data.get('row_index', '0')).strip() or '0')
        new_tag = str(data.get('tag', '')).strip()

        if sheet_name == '' or sheet_name not in REQUIRED_SHEETS:
            return jsonify({'success': False, 'message': 'Ongeldige sheet-naam'}), 400

        if row_index < 2:
            return jsonify({'success': False, 'message': 'Ongeldige rij-index'}), 400

        if not new_tag:
            return jsonify({'success': False, 'message': 'Tag is verplicht'}), 400

        # Optioneel: valideer dat de tag uit de lijst komt
        if TAGS and new_tag not in TAGS:
            return jsonify({'success': False, 'message': 'Tag is niet toegestaan'}), 400

        wb = load_workbook(EXCEL_FILE_PATH)
        if sheet_name not in wb.sheetnames:
            return jsonify({'success': False, 'message': 'Sheet niet gevonden in Excel bestand'}), 400
        sheet = wb[sheet_name]

        # Schrijf tag in kolom 12 (Tag)
        sheet.cell(row=row_index, column=12, value=new_tag)
        wb.save(EXCEL_FILE_PATH)

        user = getpass.getuser()
        logging.info(f"TAG BIJGEWERKT | Gebruiker: {user} | Sheet: {sheet_name} | Rij: {row_index} | Tag: {new_tag}")

        return jsonify({'success': True, 'message': 'Tag bijgewerkt'})
    except Exception as e:
        logging.error(f"Fout bij bijwerken tag: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500

@app.route('/add_transaction', methods=['POST'])
def add_transaction():
    """Voeg een nieuwe transactie toe"""
    try:
        # Controleer of Excel bestand is ingesteld en bestaat
        if not EXCEL_FILE_PATH or EXCEL_FILE_PATH.strip() == "":
            return jsonify({
                'success': False, 
                'message': 'Geen Excel bestand ingesteld. Ga naar Instellingen in de hoofdapp en selecteer/upload een Excel bestand.'
            }), 400
        
        if not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({
                'success': False, 
                'message': 'Excel bestand niet gevonden. Ga naar Instellingen in de hoofdapp en selecteer/upload een geldig Excel bestand.'
            }), 400

        # Haal gegevens op uit het formulier
        data = {
            'datum': request.form.get('datum'),
            'mededelingen': request.form.get('mededelingen', ''),
            'rekening': request.form.get('rekening', ''),
            'tegenrekening': request.form.get('tegenrekening', ''),
            'code': request.form.get('code', ''),
            'af_bij': request.form.get('af_bij'),
            'bedrag': request.form.get('bedrag'),
            'mutatiesoort': request.form.get('mutatiesoort', 'Kas'),
            'saldo': request.form.get('saldo', ''),
            'tag': request.form.get('tag', '')
        }
        
        # Validatie
        if not data['datum']:
            return jsonify({'success': False, 'message': 'Datum is verplicht'}), 400
        
        if not data['mededelingen'].strip():
            return jsonify({'success': False, 'message': 'Mededeling is verplicht'}), 400
        
        if not data['bedrag'].strip():
            return jsonify({'success': False, 'message': 'Bedrag is verplicht'}), 400
        
        # Converteer bedrag (accepteer komma als decimaal scheidingsteken)
        try:
            bedrag = float(data['bedrag'].replace(',', '.'))
        except ValueError:
            return jsonify({'success': False, 'message': 'Ongeldig bedrag'}), 400
        
        # Parse datum
        try:
            datum = datetime.strptime(data['datum'], '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': 'Ongeldige datum'}), 400
        
        # Laad of maak Excel bestand
        if os.path.exists(EXCEL_FILE_PATH):
            wb = load_workbook(EXCEL_FILE_PATH)
            if EXCEL_SHEET_NAME in wb.sheetnames:
                sheet = wb[EXCEL_SHEET_NAME]
            else:
                sheet = wb.create_sheet(EXCEL_SHEET_NAME)
                # Voeg headers toe aan de nieuwe sheet
                sheet.append(REQUIRED_HEADERS)
        else:
            wb = Workbook()
            # Maak alle vereiste sheets aan met headers
            main_sheet = wb.active
            main_sheet.title = "Bankrekening"
            main_sheet.append(REQUIRED_HEADERS)
            for name in ["Spaarrekening 1", "Spaarrekening 2"]:
                s = wb.create_sheet(name)
                s.append(REQUIRED_HEADERS)
            # Selecteer de juiste sheet om te schrijven
            sheet = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME in wb.sheetnames else wb["Bankrekening"]
        
        # Voeg lege rij in op positie 2
        sheet.insert_rows(2)
        
        # Voeg data toe op rij 2
        row_data = [
            datum,
            data['mededelingen'],
            data['rekening'],
            data['tegenrekening'],
            data['code'],
            data['af_bij'],
            bedrag,
            data['mutatiesoort'],
            data['mededelingen'],
            data['saldo'],
            '',
            data['tag']
        ]
        
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=2, column=col, value=value)
        
        # Sla op
        wb.save(EXCEL_FILE_PATH)
        
        # Log de actie met meer details
        user = getpass.getuser()  # Krijg Windows username
        ip_addr = request.remote_addr  # IP adres
        logging.info(f"TRANSACTIE TOEGEVOEGD | Gebruiker: {user} | IP: {ip_addr} | Datum: {data['datum']} | "
                    f"Beschrijving: {data['mededelingen']} | Bedrag: €{bedrag} | Af/Bij: {data['af_bij']} | Tag: {data['tag']}")
        
        # Bereken nieuw totaal
        new_total = calculate_total_amount()
        
        return jsonify({
            'success': True, 
            'message': 'Transactie succesvol opgeslagen!',
            'new_total': new_total
        })
        
    except Exception as e:
        logging.error(f"Fout bij toevoegen transactie: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500

@app.route('/get_total')
def get_total():
    """Haal het huidige totaal op"""
    total = calculate_total_amount()
    return jsonify({'total': total})

@app.route('/get_transactions')
def get_transactions():
    """Haal recente transacties op (AJAX)"""
    transactions = get_recent_transactions()
    return jsonify({'transactions': transactions})

@app.route('/api/all_transactions')
def api_all_transactions():
    """Haal alle transacties op (AJAX) voor de history"""
    transactions = get_all_transactions()
    return jsonify({'transactions': transactions})

@app.route('/backup')
def backup():
    """Maak handmatig een backup"""
    success = create_backup()
    if success:
        return jsonify({'success': True, 'message': 'Backup succesvol gemaakt'})
    else:
        return jsonify({'success': False, 'message': 'Fout bij maken backup'}), 500

@app.route('/quit', methods=['POST'])
def quit_application():
    """Stop de applicatie en log dit"""
    try:
        user = getpass.getuser()
        ip_addr = request.remote_addr
        duration = request.get_json().get('duration', 'Onbekend') if request.is_json else 'Onbekend'
        
        logging.info(f"APPLICATIE AFGESLOTEN | Gebruiker: {user} | IP: {ip_addr} | Sessieduur: {duration}")
        logging.info("=" * 70)
        
        # Stuur succes response terug naar client
        response = jsonify({'success': True, 'message': 'Applicatie sluit af'})
        
        # Schedule de shutdown na een korte vertraging zodat response kan worden verzonden
        def shutdown_server():
            import time
            time.sleep(1)  # Wacht 1 seconde zodat response verzonden kan worden
            logging.info("Flask server wordt beëindigd...")
            os._exit(0)
        
        import threading
        shutdown_thread = threading.Thread(target=shutdown_server, daemon=True)
        shutdown_thread.start()
        
        return response, 200
    except Exception as e:
        logging.error(f"Fout bij afsluiten applicatie: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500

@app.route('/settings')
def settings():
    if MAIN_APP_URL:
        return redirect(f"{MAIN_APP_URL}/settings")
    return settings_locked_response()

if __name__ == '__main__':
    print("=" * 60)
    print(">> Bankrekening Debutade Web Applicatie - Startup")
    print("=" * 60)
    
    # Zorg dat log directory bestaat voordat we logging configureren
    if not os.path.exists(LOG_DIRECTORY):
        try:
            os.makedirs(LOG_DIRECTORY)
            print(f"Log directory aangemaakt: {LOG_DIRECTORY}")
        except Exception as e:
            print(f"FOUT: Kan log directory niet aanmaken: {LOG_DIRECTORY}")
            print(f"Details: {str(e)}")
            exit(1)
    
    # Configureer logging EERST zodat alle logs worden geschreven
    log_file_path = os.path.join(LOG_DIRECTORY, "bankrekening_webapp_log.txt")
    logging.basicConfig(
        filename=log_file_path,
        level=getattr(logging, LOG_LEVEL.upper(), logging.INFO),
        format="%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    
    # Valideer configuratie (maakt directories aan)
    if not validate_config():
        print("\n>> FOUT: Applicatie kan niet starten. Zorg dat config.json correct is ingesteld.")
        exit(1)
    
    # Maak backup bij starten
    create_backup()
    
    # Log startup met gebruikersinfo
    user = getpass.getuser()
    logging.info("=" * 70)
    logging.info(f"APPLICATIE GESTART | Gebruiker: {user} | Tijd: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logging.info("=" * 70)
    
    print("\n>> Applicatie is klaar om te starten!")
    print("=" * 60)
    
    # Start de Flask applicatie
    # Debug=False voor productie, True voor ontwikkeling
    # host='0.0.0.0' maakt de app toegankelijk van alle apparaten op het netwerk
    port = int(os.getenv("DEBUTADE_APP_PORT", "5004"))
    app.run(debug=False, host='127.0.0.1', port=port)
