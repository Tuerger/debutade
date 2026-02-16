"""
Voeg Bon Toe - Web Applicatie
==============================

Web applicatie voor het toevoegen van bon URLs aan kasboek en bankrekening records.

Functionaliteiten:
- Tonen van alle records uit kasboek en bankrekening Excel bestanden (alle tabs)
- Per record: knop "Bon" voor selectie SharePoint bestand en knop "Bewaar" om URL op te slaan
- Alleen PDF en JPG/JPEG bestanden toegestaan
- URL wordt opgeslagen in kolom "Bon" van het Excel bestand

Versie: 1.0
Datum: 2026-01-22
Auteur: Eric G.
"""

from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
import threading
import os
import json
import logging
import sys
from datetime import datetime

# Fix encoding voor Windows console
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Zorg dat server TCP sockets onmiddellijk kan hergebruiken
app.config['ENV'] = 'production'

# Bepaal het directory waar het script zich bevindt
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.getenv(
    "DEBUTADE_CONFIG",
    os.path.abspath(os.path.join(SCRIPT_DIR, "..", "config.json")),
)

# Laad configuratie
def load_config():
    """Laad configuratie uit config.json"""
    if not os.path.exists(CONFIG_PATH):
        raise FileNotFoundError(f"Configuratiebestand niet gevonden: {CONFIG_PATH}")

    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        root_config = json.load(f)

    if "bontoevoegen" not in root_config:
        raise KeyError("Configuratiesectie ontbreekt: bontoevoegen")

    config = root_config["bontoevoegen"]
    shared = root_config.get("shared", {})
    for key in ("backup_directory", "log_directory", "log_level"):
        if key in shared:
            config[key] = shared[key]
    if shared.get("grootboek_directory"):
        if config.get("bank_excel_file_name"):
            config["bank_excel_file_path"] = os.path.join(
                shared["grootboek_directory"],
                config["bank_excel_file_name"],
            )
        if config.get("kas_excel_file_name"):
            config["kas_excel_file_path"] = os.path.join(
                shared["grootboek_directory"],
                config["kas_excel_file_name"],
            )

    required_keys = ["bank_excel_file_name", "kas_excel_file_name"]
    for key in required_keys:
        if key not in config:
            raise KeyError(f"Configuratiesleutel ontbreekt: {key}")

    return config

# Laad configuratie
config = load_config()
BANK_EXCEL_PATH = config["bank_excel_file_path"]
KAS_EXCEL_PATH = config["kas_excel_file_path"]
BACKUP_DIRECTORY = config.get("backup_directory", os.path.join(SCRIPT_DIR, "backup"))
LOG_DIRECTORY = config.get("log_directory", os.path.join(SCRIPT_DIR, "logs"))
LOG_LEVEL = config.get("log_level", "INFO")
SHAREPOINT_TENANT = config.get("sharepoint_tenant", "")
MAIN_APP_URL = os.getenv("MAIN_APP_URL", "").strip()

# Setup logging
if not os.path.exists(LOG_DIRECTORY):
    os.makedirs(LOG_DIRECTORY)

# Maak backup directory aan als deze niet bestaat
if not os.path.exists(BACKUP_DIRECTORY):
    os.makedirs(BACKUP_DIRECTORY)

log_file = os.path.join(LOG_DIRECTORY, f"voegbontoe_{datetime.now().strftime('%Y%m%d')}.log")
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL),
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# Houd workbooks in geheugen voor snellere reads/writes
workbook_cache = {}

def create_backup(file_path):
    """
    Maak een backup van het Excel bestand met timestamp.
    Returns: (success: bool, backup_path: str)
    """
    try:
        if not os.path.exists(file_path):
            logging.warning(f"Kan geen backup maken: bestand bestaat niet: {file_path}")
            return False, None
        
        # Haal bestandsnaam op
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        
        # Maak backup filename met timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f"{name}_backup_{timestamp}{ext}"
        backup_path = os.path.join(BACKUP_DIRECTORY, backup_filename)
        
        # Kopieer bestand
        from shutil import copy2
        copy2(file_path, backup_path)
        
        logging.info(f"Backup gemaakt: {backup_path}")
        return True, backup_path
    
    except Exception as e:
        logging.error(f"Fout bij maken backup: {str(e)}")
        return False, None


def _get_workbook_entry(file_path):
    """Haal (en cache) een workbook plus lock. Herlaadt bij gewijzigde mtime."""
    if not os.path.exists(file_path):
        return None

    current_mtime = os.path.getmtime(file_path)
    entry = workbook_cache.get(file_path)

    needs_reload = (
        entry is None or
        entry.get('mtime') != current_mtime or
        entry.get('wb') is None
    )

    if needs_reload:
        wb = load_workbook(file_path, data_only=True)
        entry = {
            'wb': wb,
            'mtime': current_mtime,
            'lock': threading.Lock()
        }
        workbook_cache[file_path] = entry
        logging.debug(f"Workbook geladen en gecached: {file_path}")

    return entry

def read_excel_all_tabs(file_path):
    """
    Lees alle tabs van een Excel bestand en retourneer records.
    Retourneert een lijst met dictionaries met keys: 
    tab, row_index (1-based Excel row), en kolom data
    """
    entry = _get_workbook_entry(file_path)
    if entry is None:
        logging.error(f"Excel bestand niet gevonden: {file_path}")
        return []

    records = []
    try:
        with entry['lock']:
            wb = entry['wb']
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Lees headers (eerste rij)
                headers = []
                for cell in sheet[1]:
                    headers.append(cell.value if cell.value else "")
                
                # Lees data rijen (vanaf rij 2)
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    record = {
                        'tab': sheet_name,
                        'row_index': row_idx,
                        'file_path': file_path
                    }
                    
                    # Voeg kolom data toe
                    for col_idx, (header, value) in enumerate(zip(headers, row)):
                        record[header] = value
                    
                    records.append(record)

        logging.info(f"Gelezen {len(records)} records uit cache: {file_path}")
    except Exception as e:
        logging.error(f"Fout bij lezen Excel bestand {file_path}: {str(e)}")
    
    return records

def save_bon_url_to_excel(file_path, tab_name, row_index, bon_url):
    """
    Sla de bon URL op in de kolom 'Bon' van het Excel bestand.
    Returns: (success: bool, message: str)
    """
    try:
        logging.info(f"save_bon_url_to_excel: Start - file={file_path}, tab={tab_name}, row={row_index}")
        
        # Check of bestand bestaat
        if not os.path.exists(file_path):
            logging.error(f"Bestand niet gevonden: {file_path}")
            return False, f"Excel bestand niet gevonden: {file_path}"

        logging.info("Bestand bestaat, start laden workbook...")
        # Laad workbook DIRECT (niet uit cache) voor betrouwbaarheid op OneDrive
        # Gebruik keep_vba=False en data_only=False voor snellere load
        wb = load_workbook(file_path, keep_vba=False, data_only=False)
        logging.info("Workbook geladen")
        
        if tab_name not in wb.sheetnames:
            logging.error(f"Tab niet gevonden: {tab_name}")
            wb.close()
            return False, f"Tab '{tab_name}' niet gevonden in Excel bestand"
        
        sheet = wb[tab_name]
        logging.info(f"Sheet '{tab_name}' geselecteerd")
        
        # Zoek de kolom index voor "Bon"
        bon_col_idx = None
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value and str(cell.value).strip().lower() == "bon":
                bon_col_idx = col_idx
                break
        
        if bon_col_idx is None:
            logging.error("Kolom 'Bon' niet gevonden")
            wb.close()
            return False, "Kolom 'Bon' niet gevonden in Excel bestand"
        
        logging.info(f"Kolom 'Bon' gevonden op index {bon_col_idx}")
        
        # Schrijf URL naar cel
        sheet.cell(row=row_index, column=bon_col_idx, value=bon_url)
        logging.info("URL geschreven naar cel, start opslaan...")
        
        # Sla direct op (geen cache)
        wb.save(file_path)
        logging.info("✓ Workbook opgeslagen!")
        
        # Probeer workbook te sluiten - maar laat het proces niet hangen als OneDrive een lock heeft
        try:
            wb.close()
            logging.info("Workbook gesloten")
        except Exception as close_error:
            logging.warning(f"Kon workbook niet sluiten (OneDrive lock?): {close_error} - maar opslaan is gelukt!")
        
        # Invalideer cache (veilig)
        try:
            if file_path in workbook_cache:
                del workbook_cache[file_path]
                logging.info("Cache geïnvalideerd")
        except Exception as cache_error:
            logging.warning(f"Kon cache niet invalideren: {cache_error} - niet kritiek")
        
        logging.info(f"✓✓✓ BON URL SUCCESVOL OPGESLAGEN: {file_path}, tab={tab_name}, rij={row_index}")
        return True, "Bon URL succesvol opgeslagen"
    
    except Exception as e:
        logging.error(f"FOUT bij opslaan bon URL: {str(e)}", exc_info=True)
        return False, f"Fout bij opslaan: {str(e)}"


@app.route('/')
def index():
    """Hoofdpagina - toon alle records"""
    # Lees records uit beide Excel bestanden
    kas_records = read_excel_all_tabs(KAS_EXCEL_PATH)
    bank_records = read_excel_all_tabs(BANK_EXCEL_PATH)
    
    # Markeer bron
    for record in kas_records:
        record['bron'] = 'Kasboek'
    for record in bank_records:
        record['bron'] = 'Bankrekening'
    
    # Combineer records
    all_records = kas_records + bank_records
    
    # Sorteer records van nieuw naar oud op basis van Datum kolom
    # Als er geen Datum kolom is, sorteer dan op row_index (nieuwere rijen hebben hogere nummers)
    def get_sort_key(record):
        datum = record.get('Datum') or record.get('datum') or record.get('DATUM')
        if datum and isinstance(datum, datetime):
            return datum
        elif datum:
            # Probeer datum te parsen als string
            try:
                from datetime import datetime as dt
                # Probeer verschillende datum formaten
                for fmt in ['%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d']:
                    try:
                        return dt.strptime(str(datum), fmt)
                    except:
                        continue
            except:
                pass
        # Fallback: gebruik row_index en tab als secundaire sortering
        # Hogere row_index = nieuwer, dus negatief voor reverse sort
        return datetime(1900, 1, 1)  # Oude datum voor records zonder datum
    
    # Sorteer: nieuwste eerst (reverse=True)
    all_records.sort(key=get_sort_key, reverse=True)
    
    # Bereken statistieken per tab
    stats_by_tab = {}
    
    for record in all_records:
        tab = record['tab']
        has_bon = bool(record.get('Bon'))
        
        # Statistieken per tab
        if tab not in stats_by_tab:
            stats_by_tab[tab] = {
                'total': 0,
                'with_bon': 0,
                'without_bon': 0
            }
        
        stats_by_tab[tab]['total'] += 1
        if has_bon:
            stats_by_tab[tab]['with_bon'] += 1
        else:
            stats_by_tab[tab]['without_bon'] += 1
    
    # Haal huidige datum en gebruiker op
    current_date = datetime.now().strftime('%d-%m-%Y')
    current_user = os.getlogin()
    
    return render_template('voegbontoe.html', 
                         records=all_records,
                         current_date=current_date,
                         current_user=current_user,
                         stats_by_tab=stats_by_tab,
                         main_app_url=MAIN_APP_URL)

@app.route('/save_bon_url', methods=['POST'])
def save_bon_url():
    """API endpoint om bon URL op te slaan"""
    try:
        logging.info("=== save_bon_url endpoint aangeroepen ===")
        data = request.json
        file_path = data.get('file_path')
        tab_name = data.get('tab')
        row_index = data.get('row_index')
        bon_url = data.get('bon_url')
        
        logging.info(f"Data ontvangen: file={file_path}, tab={tab_name}, row={row_index}")
        
        if not all([file_path, tab_name, row_index, bon_url]):
            logging.warning("Ontbrekende gegevens in request")
            return jsonify({'success': False, 'message': 'Ontbrekende gegevens'}), 400
        
        # Valideer SharePoint URL formaat
        if not bon_url.startswith('https://'):
            logging.warning("URL begint niet met https://")
            return jsonify({
                'success': False,
                'message': 'URL moet beginnen met https://'
            }), 400
        
        # Check of het een SharePoint URL is (moet sharepoint.com bevatten)
        if 'sharepoint.com' not in bon_url.lower():
            logging.warning("Geen SharePoint URL (mist sharepoint.com)")
            return jsonify({
                'success': False, 
                'message': 'Alleen SharePoint URLs zijn toegestaan (moet sharepoint.com bevatten)'
            }), 400
        
        # Valideer dat de URL een tenant naam heeft (formaat: https://tenant.sharepoint.com of https://tenant-my.sharepoint.com)
        import re
        sharepoint_pattern = r'https://[a-zA-Z0-9\-]+\.sharepoint\.com/'
        if not re.match(sharepoint_pattern, bon_url):
            logging.warning("Ongeldig SharePoint URL formaat")
            return jsonify({
                'success': False,
                'message': 'Ongeldig SharePoint URL formaat (verwacht: https://tenant.sharepoint.com/...)'
            }), 400
        
        # Valideer tenant naam als deze is geconfigureerd
        if SHAREPOINT_TENANT:
            tenant_pattern = f'https://{SHAREPOINT_TENANT}(-my)?\.sharepoint\.com/'
            if not re.match(tenant_pattern, bon_url, re.IGNORECASE):
                logging.warning(f"URL is niet van de juiste tenant (verwacht: {SHAREPOINT_TENANT})")
                return jsonify({
                    'success': False,
                    'message': f'URL moet van tenant "{SHAREPOINT_TENANT}" zijn (bijv. https://{SHAREPOINT_TENANT}.sharepoint.com/...)'
                }), 400
        
        logging.info("URL validatie succesvol, start opslaan naar Excel...")
        # Sla SYNCHROON op (zoals bankrekening app) voor directe feedback
        success, message = save_bon_url_to_excel(
            file_path,
            tab_name,
            row_index,
            bon_url
        )
        logging.info(f"Opslaan voltooid: success={success}, message={message}")
        
        if success:
            logging.info(f"=== SUCCES - Stuur response naar browser ===")
            response = jsonify({
                'success': True,
                'message': 'Bon URL succesvol opgeslagen'
            })
            logging.info(f"Response aangemaakt, versturen...")
            return response, 200
        else:
            logging.error(f"FOUT bij opslaan bon URL: {message}")
            return jsonify({
                'success': False,
                'message': message
            }), 500
    
    except Exception as e:
        logging.error(f"Fout in save_bon_url endpoint: {str(e)}")
        return jsonify({'success': False, 'message': f'Serverfout: {str(e)}'}), 500

@app.route('/quit', methods=['POST'])
def quit_app():
    """Sluit de applicatie af"""
    try:
        logging.info("Applicatie wordt afgesloten...")
        logging.info("=" * 70)
        
        # Stuur succes response terug naar client
        response = jsonify({'success': True, 'message': 'Applicatie sluit af'})
        
        # Schedule de shutdown na een korte vertraging zodat response kan worden verzonden
        def shutdown_server():
            import time
            time.sleep(1)  # Wacht 1 seconde zodat response verzonden kan worden
            logging.info("Flask server wordt beëindigd...")
            os._exit(0)
        
        import threading
        shutdown_thread = threading.Thread(target=shutdown_server, daemon=True)
        shutdown_thread.start()
        
        return response, 200
    except Exception as e:
        logging.error(f"Fout bij afsluiten applicatie: {str(e)}")
        return jsonify({'success': False, 'message': f'Fout: {str(e)}'}), 500

if __name__ == '__main__':
    logging.info("=" * 70)
    logging.info("VOEG BON TOE - START")
    logging.info("=" * 70)
    logging.info(f"Kasboek bestand: {KAS_EXCEL_PATH}")
    logging.info(f"Bank bestand: {BANK_EXCEL_PATH}")
    logging.info(f"Backup directory: {BACKUP_DIRECTORY}")
    logging.info("=" * 70)
    
    # Controleer of bestanden bestaan
    kas_exists = os.path.exists(KAS_EXCEL_PATH)
    bank_exists = os.path.exists(BANK_EXCEL_PATH)
    
    if not kas_exists:
        logging.warning(f"Kasboek bestand niet gevonden: {KAS_EXCEL_PATH}")
    if not bank_exists:
        logging.warning(f"Bank bestand niet gevonden: {BANK_EXCEL_PATH}")
    
    # Maak backups van beide bestanden bij opstarten
    if kas_exists or bank_exists:
        logging.info("=" * 70)
        logging.info("BACKUPS MAKEN...")
        logging.info("=" * 70)
        
        if kas_exists:
            success, backup_path = create_backup(KAS_EXCEL_PATH)
            if success:
                logging.info(f"✓ Kasboek backup: {backup_path}")
            else:
                logging.warning("✗ Kasboek backup mislukt")
        
        if bank_exists:
            success, backup_path = create_backup(BANK_EXCEL_PATH)
            if success:
                logging.info(f"✓ Bank backup: {backup_path}")
            else:
                logging.warning("✗ Bank backup mislukt")
        
        logging.info("=" * 70)

    # Preload workbooks in cache voor snellere toegang
    if kas_exists:
        _get_workbook_entry(KAS_EXCEL_PATH)
    if bank_exists:
        _get_workbook_entry(BANK_EXCEL_PATH)
    
    # Server starten met verbeterde error handling
    port = int(os.getenv("DEBUTADE_APP_PORT", "5004"))
    try:
        app.run(debug=False, host='127.0.0.1', port=port, use_reloader=False, threaded=True)
    except OSError as e:
        if "Address already in use" in str(e) or "port is in use" in str(e).lower():
            logging.error("FOUT: Poort %s is al in gebruik door een ander proces!", port)
            logging.error("Mogelijke oplossingen:")
            logging.error("1. Sluit alle voige 'Voeg Bon Toe' vensters")
            logging.error("2. Herstart uw computer")
            logging.error("3. Voer in PowerShell uit: netstat -ano | findstr :%s", port)
        else:
            logging.error(f"Kan server niet starten: {str(e)}")
        sys.exit(1)
